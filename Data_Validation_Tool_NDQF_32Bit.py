import re, os, sys, itertools, string, time
import numpy as np
import pandas as pd
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QTimer, Qt
from PyQt5.QtWidgets import QFileDialog, QLabel,QMessageBox, QSizePolicy, QTabWidget, QWidget
import matplotlib.pyplot as plt
from matplotlib.pyplot import axes, figure
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from HealthSubCentreVal import *
from PrimaryHealthCentre import *
from SubDistrictHospitalVal import *
from DistrictHospitalVal import *
from CommunityHealthCentreVal import *
from pandas.io.formats import style
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from PyQt5.QtCore import QThread, pyqtSignal, QObject, pyqtSlot
from PyQt5.QtWidgets import QWidget, QProgressBar, QVBoxLayout


## SCROLLABLE MENU FOR FILTERS
##############################

class ScrollableMenu(QtWidgets.QMenu):
    deltaY = 0
    dirty = True
    ignoreAutoScroll = False
    def __init__(self, *args, **kwargs):
        maxItemCount = kwargs.pop('maxItemCount', 0)
        super().__init__(*args, **kwargs)
        self._maximumHeight = self.maximumHeight()
        self._actionRects = []

        self.scrollTimer = QtCore.QTimer(self, interval=50, singleShot=True, timeout=self.checkScroll)
        self.scrollTimer.setProperty('defaultInterval', 50)
        self.delayTimer = QtCore.QTimer(self, interval=100, singleShot=True)

        self.setMaxItemCount(maxItemCount)

    @property
    def actionRects(self):
        if self.dirty or not self._actionRects:
            self._actionRects.clear()
            offset = self.offset()
            for action in self.actions():
                geo = super().actionGeometry(action)
                if offset:
                    geo.moveTop(geo.y() - offset)
                self._actionRects.append(geo)
            self.dirty = False
        return self._actionRects

    def iterActionRects(self):
        for action, rect in zip(self.actions(), self.actionRects):
            yield action, rect

    def setMaxItemCount(self, count):
        style = self.style()
        opt = QtWidgets.QStyleOptionMenuItem()
        opt.initFrom(self)

        a = QtWidgets.QAction('fake action', self)
        self.initStyleOption(opt, a)
        size = QtCore.QSize()
        fm = self.fontMetrics()
        qfm = opt.fontMetrics
        size.setWidth(fm.boundingRect(QtCore.QRect(), QtCore.Qt.TextSingleLine, a.text()).width())
        size.setHeight(max(fm.height(), qfm.height()))
        self.defaultItemHeight = style.sizeFromContents(style.CT_MenuItem, opt, size, self).height()

        if not count:
            self.setMaximumHeight(self._maximumHeight)
        else:
            fw = style.pixelMetric(style.PM_MenuPanelWidth, None, self)
            vmargin = style.pixelMetric(style.PM_MenuHMargin, opt, self)
            scrollHeight = self.scrollHeight(style)
            self.setMaximumHeight(
                self.defaultItemHeight * count + (fw + vmargin + scrollHeight) * 2)
        self.dirty = True

    def scrollHeight(self, style):
        return style.pixelMetric(style.PM_MenuScrollerHeight, None, self) * 2

    def isScrollable(self):
        return self.height() < super().sizeHint().height()

    def checkScroll(self):
        pos = self.mapFromGlobal(QtGui.QCursor.pos())
        delta = max(2, int(self.defaultItemHeight * .25))
        if pos in self.scrollUpRect:
            delta *= -1
        elif pos not in self.scrollDownRect:
            return
        if self.scrollBy(delta):
            self.scrollTimer.start(self.scrollTimer.property('defaultInterval'))

    def offset(self):
        if self.isScrollable():
            return self.deltaY - self.scrollHeight(self.style())
        return 0

    def translatedActionGeometry(self, action):
        return self.actionRects[self.actions().index(action)]

    def ensureVisible(self, action):
        style = self.style()
        fw = style.pixelMetric(style.PM_MenuPanelWidth, None, self)
        hmargin = style.pixelMetric(style.PM_MenuHMargin, None, self)
        vmargin = style.pixelMetric(style.PM_MenuVMargin, None, self)
        scrollHeight = self.scrollHeight(style)
        extent = fw + hmargin + vmargin + scrollHeight
        r = self.rect().adjusted(0, extent, 0, -extent)
        geo = self.translatedActionGeometry(action)
        if geo.top() < r.top():
            self.scrollBy(-(r.top() - geo.top()))
        elif geo.bottom() > r.bottom():
            self.scrollBy(geo.bottom() - r.bottom())

    def scrollBy(self, step):
        if step < 0:
            newDelta = max(0, self.deltaY + step)
            if newDelta == self.deltaY:
                return False
        elif step > 0:
            newDelta = self.deltaY + step
            style = self.style()
            scrollHeight = self.scrollHeight(style)
            bottom = self.height() - scrollHeight

            for lastAction in reversed(self.actions()):
                if lastAction.isVisible():
                    break
            lastBottom = self.actionGeometry(lastAction).bottom() - newDelta + scrollHeight
            if lastBottom < bottom:
                newDelta -= bottom - lastBottom
            if newDelta == self.deltaY:
                return False

        self.deltaY = newDelta
        self.dirty = True
        self.update()
        return True

    def actionAt(self, pos):
        for action, rect in self.iterActionRects():
            if pos in rect:
                return action

    # class methods reimplementation

    def sizeHint(self):
        hint = super().sizeHint()
        if hint.height() > self.maximumHeight():
            hint.setHeight(self.maximumHeight())
        return hint

    def eventFilter(self, source, event):
        if event.type() == event.Show:
            if self.isScrollable() and self.deltaY:
                action = source.menuAction()
                self.ensureVisible(action)
                rect = self.translatedActionGeometry(action)
                delta = rect.topLeft() - self.actionGeometry(action).topLeft()
                source.move(source.pos() + delta)
            return False
        return super().eventFilter(source, event)

    def event(self, event):
        if not self.isScrollable():
            return super().event(event)
        if event.type() == event.KeyPress and event.key() in (QtCore.Qt.Key_Up, QtCore.Qt.Key_Down):
            res = super().event(event)
            action = self.activeAction()
            if action:
                self.ensureVisible(action)
                self.update()
            return res
        elif event.type() in (event.MouseButtonPress, event.MouseButtonDblClick):
            pos = event.pos()
            if pos in self.scrollUpRect or pos in self.scrollDownRect:
                if event.button() == QtCore.Qt.LeftButton:
                    step = max(2, int(self.defaultItemHeight * .25))
                    if pos in self.scrollUpRect:
                        step *= -1
                    self.scrollBy(step)
                    self.scrollTimer.start(200)
                    self.ignoreAutoScroll = True
                return True
        elif event.type() == event.MouseButtonRelease:
            pos = event.pos()
            self.scrollTimer.stop()
            if not (pos in self.scrollUpRect or pos in self.scrollDownRect):
                action = self.actionAt(event.pos())
                if action:
                    action.trigger()
                    self.close()
            return True
        return super().event(event)

    def timerEvent(self, event):
        if not self.isScrollable():
            # ignore internal timer event for reopening popups
            super().timerEvent(event)

    def mouseMoveEvent(self, event):
        if not self.isScrollable():
            super().mouseMoveEvent(event)
            return

        pos = event.pos()
        if pos.y() < self.scrollUpRect.bottom() or pos.y() > self.scrollDownRect.top():
            if not self.ignoreAutoScroll and not self.scrollTimer.isActive():
                self.scrollTimer.start(200)
            return
        self.ignoreAutoScroll = False

        oldAction = self.activeAction()
        if not pos in self.rect():
            action = None
        else:
            y = event.y()
            for action, rect in self.iterActionRects():
                if rect.y() <= y <= rect.y() + rect.height():
                    break
            else:
                action = None

        self.setActiveAction(action)
        if action and not action.isSeparator():
            def ensureVisible():
                self.delayTimer.timeout.disconnect()
                self.ensureVisible(action)
            try:
                self.delayTimer.disconnect()
            except:
                pass
            self.delayTimer.timeout.connect(ensureVisible)
            self.delayTimer.start(150)
        elif oldAction and oldAction.menu() and oldAction.menu().isVisible():
            def closeMenu():
                self.delayTimer.timeout.disconnect()
                oldAction.menu().hide()
            self.delayTimer.timeout.connect(closeMenu)
            self.delayTimer.start(50)
        self.update()

    def wheelEvent(self, event):
        if not self.isScrollable():
            return
        self.delayTimer.stop()
        if event.angleDelta().y() < 0:
            self.scrollBy(self.defaultItemHeight)
        else:
            self.scrollBy(-self.defaultItemHeight)

    def showEvent(self, event):
        if self.isScrollable():
            self.deltaY = 0
            self.dirty = True
            for action in self.actions():
                if action.menu():
                    action.menu().installEventFilter(self)
            self.ignoreAutoScroll = False
        super().showEvent(event)

    def hideEvent(self, event):
        for action in self.actions():
            if action.menu():
                action.menu().removeEventFilter(self)
        super().hideEvent(event)

    def resizeEvent(self, event):
        super().resizeEvent(event)

        style = self.style()
        l, t, r, b = self.getContentsMargins()
        fw = style.pixelMetric(style.PM_MenuPanelWidth, None, self)
        hmargin = style.pixelMetric(style.PM_MenuHMargin, None, self)
        vmargin = style.pixelMetric(style.PM_MenuVMargin, None, self)
        leftMargin = fw + hmargin + l
        topMargin = fw + vmargin + t
        bottomMargin = fw + vmargin + b
        contentWidth = self.width() - (fw + hmargin) * 2 - l - r

        scrollHeight = self.scrollHeight(style)
        self.scrollUpRect = QtCore.QRect(leftMargin, topMargin, contentWidth, scrollHeight)
        self.scrollDownRect = QtCore.QRect(leftMargin, self.height() - scrollHeight - bottomMargin, 
            contentWidth, scrollHeight)

    def paintEvent(self, event):
        if not self.isScrollable():
            super().paintEvent(event)
            return

        style = self.style()
        qp = QtGui.QPainter(self)
        rect = self.rect()
        emptyArea = QtGui.QRegion(rect)

        menuOpt = QtWidgets.QStyleOptionMenuItem()
        menuOpt.initFrom(self)
        menuOpt.state = style.State_None
        menuOpt.maxIconWidth = 0
        menuOpt.tabWidth = 0
        style.drawPrimitive(style.PE_PanelMenu, menuOpt, qp, self)

        fw = style.pixelMetric(style.PM_MenuPanelWidth, None, self)

        topEdge = self.scrollUpRect.bottom()
        bottomEdge = self.scrollDownRect.top()

        offset = self.offset()
        qp.save()
        qp.translate(0, -offset)
        # offset translation is required in order to allow correct fade animations
        for action, actionRect in self.iterActionRects():
            actionRect = self.translatedActionGeometry(action)
            if actionRect.bottom() < topEdge:
                continue
            if actionRect.top() > bottomEdge:
                continue

            visible = QtCore.QRect(actionRect)
            if actionRect.bottom() > bottomEdge:
                visible.setBottom(bottomEdge)
            elif actionRect.top() < topEdge:
                visible.setTop(topEdge)
            visible = QtGui.QRegion(visible.translated(0, offset))
            qp.setClipRegion(visible)
            emptyArea -= visible.translated(0, -offset)

            opt = QtWidgets.QStyleOptionMenuItem()
            self.initStyleOption(opt, action)
            opt.rect = actionRect.translated(0, offset)
            style.drawControl(style.CE_MenuItem, opt, qp, self)
        qp.restore()

        cursor = self.mapFromGlobal(QtGui.QCursor.pos())
        upData = (
            False, self.deltaY > 0, self.scrollUpRect
        )
        downData = (
            True, actionRect.bottom() - 2 > bottomEdge, self.scrollDownRect
        )

        for isDown, enabled, scrollRect in upData, downData:
            qp.setClipRect(scrollRect)

            scrollOpt = QtWidgets.QStyleOptionMenuItem()
            scrollOpt.initFrom(self)
            scrollOpt.state = style.State_None
            scrollOpt.checkType = scrollOpt.NotCheckable
            scrollOpt.maxIconWidth = scrollOpt.tabWidth = 0
            scrollOpt.rect = scrollRect
            scrollOpt.menuItemType = scrollOpt.Scroller
            if enabled:
                if cursor in scrollRect:
                    frame = QtWidgets.QStyleOptionMenuItem()
                    frame.initFrom(self)
                    frame.rect = scrollRect
                    frame.state |= style.State_Selected | style.State_Enabled
                    style.drawControl(style.CE_MenuItem, frame, qp, self)

                scrollOpt.state |= style.State_Enabled
                scrollOpt.palette.setCurrentColorGroup(QtGui.QPalette.Active)
            else:
                scrollOpt.palette.setCurrentColorGroup(QtGui.QPalette.Disabled)
            if isDown:
                scrollOpt.state |= style.State_DownArrow
            style.drawControl(style.CE_MenuScroller, scrollOpt, qp, self)

        if fw:
            borderReg = QtGui.QRegion()
            borderReg |= QtGui.QRegion(QtCore.QRect(0, 0, fw, self.height()))
            borderReg |= QtGui.QRegion(QtCore.QRect(self.width() - fw, 0, fw, self.height()))
            borderReg |= QtGui.QRegion(QtCore.QRect(0, 0, self.width(), fw))
            borderReg |= QtGui.QRegion(QtCore.QRect(0, self.height() - fw, self.width(), fw))
            qp.setClipRegion(borderReg)
            emptyArea -= borderReg
            frame = QtWidgets.QStyleOptionFrame()
            frame.rect = rect
            frame.palette = self.palette()
            frame.state = QtWidgets.QStyle.State_None
            frame.lineWidth = style.pixelMetric(style.PM_MenuPanelWidth)
            frame.midLineWidth = 0
            style.drawPrimitive(style.PE_FrameMenu, frame, qp, self)

        qp.setClipRegion(emptyArea)
        menuOpt.state = style.State_None
        menuOpt.menuItemType = menuOpt.EmptyArea
        menuOpt.checkType = menuOpt.NotCheckable
        menuOpt.rect = menuOpt.menuRect = rect
        style.drawControl(style.CE_MenuEmptyArea, menuOpt, qp, self)

## PROGRESS BAR
## ============
class PopUpProgressBar(QWidget):

    def __init__(self):
        super().__init__()
        self.pbar = QProgressBar(self)
        self.label_1 = QLabel(self)
        self.pbar.setGeometry(30, 40, 500, 75)
        self.label_1.setGeometry(40, 50, 500, 75)
        self.label_1.setText("PLEASE WAIT...!")
        self.layout = QVBoxLayout()
        self.layout.addWidget(self.label_1)
        self.layout.addWidget(self.pbar)
        self.setLayout(self.layout)
        self.setGeometry(700, 500, 500, 200)
        self.setWindowTitle('WORK IN PROGRESS, PLEASE WAIT ...!')

        # self.obj = Worker()
        # self.thread = QThread()
        # self.obj.intReady.connect(self.on_count_changed)
        # self.obj.moveToThread(self.thread)
        # self.obj.finished.connect(self.thread.quit)
        # self.obj.finished.connect(self.hide)  # To hide the progress bar after the progress is completed
        # self.thread.started.connect(self.obj.proc_counter)
        # self.thread.start()  # This was moved to start_progress

    def start_progress(self):  # To restart the progress every time
        self.show()
        # self.thread.start()


### USER INTERFACE CODE ###
### =================== ###
class Ui_TabWidget(object):
    def setupUi(self, TabWidget):
        TabWidget.setObjectName("TabWidget")
        TabWidget.resize(1587, 813)
        TabWidget.setStyleSheet("QTabWidget{background-color: rgb(255, 237, 242);}\n"
"")
        self.tab = QtWidgets.QWidget()
        self.tab.setStyleSheet("background-color: rgb(255, 255, 255);")
        self.tab.setObjectName("tab")
        self.gridLayout_2 = QtWidgets.QGridLayout(self.tab)
        self.gridLayout_2.setObjectName("gridLayout_2")
        self.gridLayout = QtWidgets.QGridLayout()
        self.gridLayout.setObjectName("gridLayout")
        self.label_7 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_7.sizePolicy().hasHeightForWidth())
        self.label_7.setSizePolicy(sizePolicy)
        self.label_7.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_7.setAlignment(QtCore.Qt.AlignCenter)
        self.label_7.setObjectName("label_7")
        self.gridLayout.addWidget(self.label_7, 8, 0, 1, 1)
        self.pushButton_11 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_11.sizePolicy().hasHeightForWidth())
        self.pushButton_11.setSizePolicy(sizePolicy)
        self.pushButton_11.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_11.setFont(font)
        self.pushButton_11.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_11.setObjectName("pushButton_11")
        self.pushButton_11.installEventFilter(TabWidget)
        self.pushButton_11.clicked.connect(self.onSelectRuralUrban)

        self.gridLayout.addWidget(self.pushButton_11, 8, 6, 1, 1)
        self.pushButton_12 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_12.sizePolicy().hasHeightForWidth())
        self.pushButton_12.setSizePolicy(sizePolicy)
        self.pushButton_12.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_12.setFont(font)
        self.pushButton_12.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_12.setObjectName("pushButton_12")
        self.pushButton_12.installEventFilter(TabWidget)
        self.pushButton_12.clicked.connect(self.onSelectOwnership)

        self.gridLayout.addWidget(self.pushButton_12, 9, 6, 1, 1)
        self.label_11 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_11.sizePolicy().hasHeightForWidth())
        self.label_11.setSizePolicy(sizePolicy)
        self.label_11.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_11.setFont(font)
        self.label_11.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_11.setAlignment(QtCore.Qt.AlignCenter)
        self.label_11.setObjectName("label_11")
        self.gridLayout.addWidget(self.label_11, 11, 0, 1, 1)
        self.label_9 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_9.sizePolicy().hasHeightForWidth())
        self.label_9.setSizePolicy(sizePolicy)
        self.label_9.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_9.setFont(font)
        self.label_9.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_9.setAlignment(QtCore.Qt.AlignCenter)
        self.label_9.setObjectName("label_9")
        self.gridLayout.addWidget(self.label_9, 9, 0, 1, 1)
        self.lineEdit_3 = QtWidgets.QLineEdit(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_3.sizePolicy().hasHeightForWidth())
        self.lineEdit_3.setSizePolicy(sizePolicy)
        self.lineEdit_3.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_3.setFont(font)
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.gridLayout.addWidget(self.lineEdit_3, 3, 4, 1, 2)
        self.pushButton_2 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_2.sizePolicy().hasHeightForWidth())
        self.pushButton_2.setSizePolicy(sizePolicy)
        self.pushButton_2.setMaximumSize(QtCore.QSize(16777215, 55))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        self.pushButton_2.setStyleSheet("QPushButton{background-color: #F47B1F;color: white;}\n"
"\n"
"")
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.guidelines)
        self.gridLayout.addWidget(self.pushButton_2, 1, 9, 1, 1)
        self.pushButton_5 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_5.sizePolicy().hasHeightForWidth())
        self.pushButton_5.setSizePolicy(sizePolicy)
        self.pushButton_5.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_5.setFont(font)
        self.pushButton_5.setStyleSheet("background-color: #003679;color: white;")
        self.pushButton_5.setObjectName("pushButton_5")
        self.pushButton_5.clicked.connect(self.VerifyFType)

        self.gridLayout.addWidget(self.pushButton_5, 3, 6, 1, 1)
        spacerItem = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem, 1, 8, 1, 1)
        self.pushButton_13 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_13.sizePolicy().hasHeightForWidth())
        self.pushButton_13.setSizePolicy(sizePolicy)
        self.pushButton_13.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_13.setFont(font)
        self.pushButton_13.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_13.setObjectName("pushButton_13")
        self.pushButton_13.installEventFilter(TabWidget)
        self.pushButton_13.clicked.connect(self.onSelectFacilityName)
        self.gridLayout.addWidget(self.pushButton_13, 11, 6, 1, 1)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_2.sizePolicy().hasHeightForWidth())
        self.lineEdit_2.setSizePolicy(sizePolicy)
        self.lineEdit_2.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.gridLayout.addWidget(self.lineEdit_2, 3, 0, 1, 4)
        self.label_13 = QtWidgets.QLabel(self.tab)
        self.label_13.setMaximumSize(QtCore.QSize(32, 32))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_13.setFont(font)
        self.label_13.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_13.setAlignment(QtCore.Qt.AlignCenter)
        self.label_13.setObjectName("label_13")
        self.gridLayout.addWidget(self.label_13, 3, 7, 1, 1)
        spacerItem1 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout.addItem(spacerItem1, 3, 8, 1, 1)
        self.label_18 = QtWidgets.QLabel(self.tab)
        self.label_18.setMaximumSize(QtCore.QSize(100, 32))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(10)
        font.setItalic(True)
        self.label_18.setFont(font)
        self.label_18.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_18.setAlignment(QtCore.Qt.AlignCenter)
        self.label_18.setObjectName("label_18")
        self.gridLayout.addWidget(self.label_18, 11, 9, 1, 1)
        self.pushButton = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton.sizePolicy().hasHeightForWidth())
        self.pushButton.setSizePolicy(sizePolicy)
        self.pushButton.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        self.pushButton.setStyleSheet("QPushButton{background-color: #73C067;color: white;}")
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.upload)

        # ProgressBar addition
        # --------------------
        self.popup = PopUpProgressBar()

        self.gridLayout.addWidget(self.pushButton, 1, 0, 1, 1)
        self.label_2 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_2.sizePolicy().hasHeightForWidth())
        self.label_2.setSizePolicy(sizePolicy)
        self.label_2.setMaximumSize(QtCore.QSize(16777215, 32))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_2.setFont(font)
        self.label_2.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_2.setAlignment(QtCore.Qt.AlignCenter)
        self.label_2.setObjectName("label_2")
        self.gridLayout.addWidget(self.label_2, 1, 1, 1, 1)
        self.label_6 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_6.sizePolicy().hasHeightForWidth())
        self.label_6.setSizePolicy(sizePolicy)
        self.label_6.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_6.setAlignment(QtCore.Qt.AlignCenter)
        self.label_6.setObjectName("label_6")
        self.gridLayout.addWidget(self.label_6, 7, 3, 1, 3)
        self.pushButton_10 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_10.sizePolicy().hasHeightForWidth())
        self.pushButton_10.setSizePolicy(sizePolicy)
        self.pushButton_10.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_10.setFont(font)
        self.pushButton_10.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_10.setObjectName("pushButton_10")
        self.pushButton_10.installEventFilter(TabWidget)
        self.pushButton_10.clicked.connect(self.onSelectHealthBlock)

        self.gridLayout.addWidget(self.pushButton_10, 7, 6, 1, 1)
        self.label_8 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_8.sizePolicy().hasHeightForWidth())
        self.label_8.setSizePolicy(sizePolicy)
        self.label_8.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_8.setFont(font)
        self.label_8.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_8.setAlignment(QtCore.Qt.AlignCenter)
        self.label_8.setObjectName("label_8")
        self.gridLayout.addWidget(self.label_8, 8, 3, 1, 3)
        self.lineEdit = QtWidgets.QLineEdit(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit.sizePolicy().hasHeightForWidth())
        self.lineEdit.setSizePolicy(sizePolicy)
        self.lineEdit.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit.setFont(font)
        self.lineEdit.setObjectName("lineEdit")
        self.gridLayout.addWidget(self.lineEdit, 1, 2, 1, 5)
        self.label_4 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_4.sizePolicy().hasHeightForWidth())
        self.label_4.setSizePolicy(sizePolicy)
        self.label_4.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setStyleSheet("QLabel{background-color: white;color: grey;}")
        self.label_4.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_4.setObjectName("label_4")
        self.gridLayout.addWidget(self.label_4, 5, 1, 2, 2)
        self.label_19 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_19.sizePolicy().hasHeightForWidth())
        self.label_19.setSizePolicy(sizePolicy)
        self.label_19.setMaximumSize(QtCore.QSize(16777215, 40))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(12)
        font.setItalic(True)
        self.label_19.setFont(font)
        self.label_19.setStyleSheet("QLabel{color : red;}")
        self.label_19.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_19.setObjectName("label_19")
        self.gridLayout.addWidget(self.label_19, 4, 6, 1, 4)
        self.pushButton_8 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_8.sizePolicy().hasHeightForWidth())
        self.pushButton_8.setSizePolicy(sizePolicy)
        self.pushButton_8.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_8.setFont(font)
        self.pushButton_8.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_8.setObjectName("pushButton_8")
        self.pushButton_8.installEventFilter(TabWidget)
        self.pushButton_8.clicked.connect(self.onSelectSubDistrict)

        self.gridLayout.addWidget(self.pushButton_8, 9, 1, 1, 2)
        self.pushButton_7 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_7.sizePolicy().hasHeightForWidth())
        self.pushButton_7.setSizePolicy(sizePolicy)
        self.pushButton_7.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_7.setFont(font)
        self.pushButton_7.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_7.setObjectName("pushButton_7")
        self.pushButton_7.installEventFilter(TabWidget)
        self.pushButton_7.clicked.connect(self.onSelectDistrict)

        self.gridLayout.addWidget(self.pushButton_7, 8, 1, 1, 2)
        self.pushButton_15 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_15.sizePolicy().hasHeightForWidth())
        self.pushButton_15.setSizePolicy(sizePolicy)
        self.pushButton_15.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_15.setFont(font)
        self.pushButton_15.setStyleSheet("background-color: #B00020;color: white;")
        self.pushButton_15.setObjectName("pushButton_15")
        self.pushButton_15.clicked.connect(self.reset)
        self.gridLayout.addWidget(self.pushButton_15, 9, 9, 1, 1)
        self.label_5 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_5.sizePolicy().hasHeightForWidth())
        self.label_5.setSizePolicy(sizePolicy)
        self.label_5.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setStyleSheet("color: grey;background-color: rgb(255, 255, 255);")
        self.label_5.setAlignment(QtCore.Qt.AlignCenter)
        self.label_5.setObjectName("label_5")
        self.gridLayout.addWidget(self.label_5, 7, 0, 1, 1)
        self.pushButton_9 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_9.sizePolicy().hasHeightForWidth())
        self.pushButton_9.setSizePolicy(sizePolicy)
        self.pushButton_9.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_9.setFont(font)
        self.pushButton_9.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_9.setObjectName("pushButton_9")
        self.pushButton_9.installEventFilter(TabWidget)
        self.pushButton_9.clicked.connect(self.onSelectBlock)

        self.gridLayout.addWidget(self.pushButton_9, 11, 1, 1, 2)
        self.pushButton_6 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_6.sizePolicy().hasHeightForWidth())
        self.pushButton_6.setSizePolicy(sizePolicy)
        self.pushButton_6.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_6.setFont(font)
        self.pushButton_6.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_6.setObjectName("pushButton_6")
        self.pushButton_6.installEventFilter(TabWidget)
        self.pushButton_6.clicked.connect(self.onSelectState)

        self.gridLayout.addWidget(self.pushButton_6, 7, 1, 1, 2)
        self.label_10 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_10.sizePolicy().hasHeightForWidth())
        self.label_10.setSizePolicy(sizePolicy)
        self.label_10.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_10.setFont(font)
        self.label_10.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_10.setAlignment(QtCore.Qt.AlignCenter)
        self.label_10.setObjectName("label_10")
        self.gridLayout.addWidget(self.label_10, 9, 3, 1, 3)
        self.label_12 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_12.sizePolicy().hasHeightForWidth())
        self.label_12.setSizePolicy(sizePolicy)
        self.label_12.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.label_12.setFont(font)
        self.label_12.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_12.setAlignment(QtCore.Qt.AlignCenter)
        self.label_12.setObjectName("label_12")
        self.gridLayout.addWidget(self.label_12, 11, 3, 1, 3)
        self.label = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label.sizePolicy().hasHeightForWidth())
        self.label.setSizePolicy(sizePolicy)
        self.label.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(28)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setStyleSheet("QLabel{background-color: #003679; color : white;}\n"
"")
        self.label.setAlignment(QtCore.Qt.AlignCenter)
        self.label.setObjectName("label")
        self.gridLayout.addWidget(self.label, 0, 0, 1, 10)
        self.label_3 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_3.sizePolicy().hasHeightForWidth())
        self.label_3.setSizePolicy(sizePolicy)
        self.label_3.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(12)
        font.setItalic(True)
        self.label_3.setFont(font)
        self.label_3.setStyleSheet("QLabel{color : red;}")
        self.label_3.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_3.setObjectName("label_3")
        self.gridLayout.addWidget(self.label_3, 2, 0, 1, 7)
        self.pushButton_4 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_4.sizePolicy().hasHeightForWidth())
        self.pushButton_4.setSizePolicy(sizePolicy)
        self.pushButton_4.setMaximumSize(QtCore.QSize(16777215, 55))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setStyleSheet("QPushButton{background-color: #F47B1F;color: white;}")
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_4.clicked.connect(self.UserManual)
        self.gridLayout.addWidget(self.pushButton_4, 3, 9, 1, 1)
        self.label_14 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_14.sizePolicy().hasHeightForWidth())
        self.label_14.setSizePolicy(sizePolicy)
        self.label_14.setMaximumSize(QtCore.QSize(100, 32))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(10)
        font.setItalic(True)
        self.label_14.setFont(font)
        self.label_14.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_14.setAlignment(QtCore.Qt.AlignCenter)
        self.label_14.setObjectName("label_14")
        self.gridLayout.addWidget(self.label_14, 5, 3, 2, 3)
        self.label_16 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_16.sizePolicy().hasHeightForWidth())
        self.label_16.setSizePolicy(sizePolicy)
        self.label_16.setMaximumSize(QtCore.QSize(32, 36))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_16.setFont(font)
        self.label_16.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_16.setAlignment(QtCore.Qt.AlignCenter)
        self.label_16.setObjectName("label_16")
        self.gridLayout.addWidget(self.label_16, 9, 8, 1, 1)
        self.pushButton_14 = QtWidgets.QPushButton(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_14.sizePolicy().hasHeightForWidth())
        self.pushButton_14.setSizePolicy(sizePolicy)
        self.pushButton_14.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_14.setFont(font)
        self.pushButton_14.setStyleSheet("background-color: #73C067;color: white;")
        self.pushButton_14.setObjectName("pushButton_14")
        self.gridLayout.addWidget(self.pushButton_14, 8, 9, 1, 1)
        self.pushButton_14.clicked.connect(self.export)
        self.label_15 = QtWidgets.QLabel(self.tab)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_15.sizePolicy().hasHeightForWidth())
        self.label_15.setSizePolicy(sizePolicy)
        self.label_15.setMaximumSize(QtCore.QSize(32, 32))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_15.setFont(font)
        self.label_15.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_15.setAlignment(QtCore.Qt.AlignCenter)
        self.label_15.setObjectName("label_15")
        self.gridLayout.addWidget(self.label_15, 8, 8, 1, 1)
        self.gridLayout.setColumnStretch(0, 1)
        self.gridLayout.setColumnStretch(2, 1)
        self.gridLayout.setColumnStretch(5, 1)
        self.gridLayout.setColumnStretch(6, 1)
        self.gridLayout.setColumnStretch(9, 1)
        self.gridLayout.setRowStretch(0, 2)
        self.gridLayout.setRowStretch(1, 1)
        self.gridLayout.setRowStretch(3, 1)
        self.gridLayout.setRowStretch(6, 1)
        self.gridLayout.setRowStretch(7, 1)
        self.gridLayout.setRowStretch(8, 1)
        self.gridLayout.setRowStretch(9, 1)
        self.gridLayout.setRowStretch(11, 1)
        self.gridLayout_2.addLayout(self.gridLayout, 0, 0, 1, 1)
        TabWidget.addTab(self.tab, "")
        self.tab1 = QtWidgets.QWidget()
        self.tab1.setObjectName("tab1")
        self.gridLayout_5 = QtWidgets.QGridLayout(self.tab1)
        self.gridLayout_5.setObjectName("gridLayout_5")
        self.gridLayout_4 = QtWidgets.QGridLayout()
        self.gridLayout_4.setObjectName("gridLayout_4")
        self.pushButton_44 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_44.sizePolicy().hasHeightForWidth())
        self.pushButton_44.setSizePolicy(sizePolicy)
        self.pushButton_44.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_44.setFont(font)
        self.pushButton_44.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_44.setObjectName("pushButton_44")
        self.pushButton_44.installEventFilter(TabWidget)
        self.pushButton_44.clicked.connect(self.onSelectState)

        self.gridLayout_4.addWidget(self.pushButton_44, 7, 1, 1, 2)
        self.label_53 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_53.sizePolicy().hasHeightForWidth())
        self.label_53.setSizePolicy(sizePolicy)
        self.label_53.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_53.setFont(font)
        self.label_53.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_53.setAlignment(QtCore.Qt.AlignCenter)
        self.label_53.setObjectName("label_53")
        self.gridLayout_4.addWidget(self.label_53, 9, 3, 1, 3)
        self.label_54 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_54.sizePolicy().hasHeightForWidth())
        self.label_54.setSizePolicy(sizePolicy)
        self.label_54.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_54.setFont(font)
        self.label_54.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_54.setAlignment(QtCore.Qt.AlignCenter)
        self.label_54.setObjectName("label_54")
        self.gridLayout_4.addWidget(self.label_54, 11, 3, 1, 3)
        self.label_55 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_55.sizePolicy().hasHeightForWidth())
        self.label_55.setSizePolicy(sizePolicy)
        self.label_55.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(28)
        font.setBold(True)
        font.setWeight(75)
        self.label_55.setFont(font)
        self.label_55.setStyleSheet("QLabel{background-color: #003679; color : white;}\n"
"")
        self.label_55.setAlignment(QtCore.Qt.AlignCenter)
        self.label_55.setObjectName("label_55")
        self.gridLayout_4.addWidget(self.label_55, 0, 0, 1, 10)
        self.label_56 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_56.sizePolicy().hasHeightForWidth())
        self.label_56.setSizePolicy(sizePolicy)
        self.label_56.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(12)
        font.setItalic(True)
        self.label_56.setFont(font)
        self.label_56.setStyleSheet("QLabel{color : red;}")
        self.label_56.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_56.setObjectName("label_56")
        self.gridLayout_4.addWidget(self.label_56, 2, 0, 1, 7)
        self.pushButton_45 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_45.sizePolicy().hasHeightForWidth())
        self.pushButton_45.setSizePolicy(sizePolicy)
        self.pushButton_45.setMaximumSize(QtCore.QSize(16777215, 55))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_45.setFont(font)
        self.pushButton_45.setStyleSheet("QPushButton{background-color: #F47B1F;color: white;}")
        self.pushButton_45.setObjectName("pushButton_45")
        self.pushButton_45.clicked.connect(self.UserManual)
        self.gridLayout_4.addWidget(self.pushButton_45, 3, 9, 1, 1)
        self.pushButton_37 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_37.sizePolicy().hasHeightForWidth())
        self.pushButton_37.setSizePolicy(sizePolicy)
        self.pushButton_37.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_37.setFont(font)
        self.pushButton_37.setStyleSheet("QPushButton{background-color: #73C067;color: white;}")
        self.pushButton_37.setObjectName("pushButton_37")
        self.pushButton_37.clicked.connect(self.upload)

        self.gridLayout_4.addWidget(self.pushButton_37, 1, 0, 1, 1)
        self.label_45 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_45.sizePolicy().hasHeightForWidth())
        self.label_45.setSizePolicy(sizePolicy)
        self.label_45.setMaximumSize(QtCore.QSize(16777215, 32))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_45.setFont(font)
        self.label_45.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_45.setAlignment(QtCore.Qt.AlignCenter)
        self.label_45.setObjectName("label_45")
        self.gridLayout_4.addWidget(self.label_45, 1, 1, 1, 1)
        self.label_46 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_46.sizePolicy().hasHeightForWidth())
        self.label_46.setSizePolicy(sizePolicy)
        self.label_46.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_46.setFont(font)
        self.label_46.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_46.setAlignment(QtCore.Qt.AlignCenter)
        self.label_46.setObjectName("label_46")
        self.gridLayout_4.addWidget(self.label_46, 7, 3, 1, 3)
        self.pushButton_38 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_38.sizePolicy().hasHeightForWidth())
        self.pushButton_38.setSizePolicy(sizePolicy)
        self.pushButton_38.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_38.setFont(font)
        self.pushButton_38.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_38.setObjectName("pushButton_38")
        self.pushButton_38.installEventFilter(TabWidget)
        self.pushButton_38.clicked.connect(self.onSelectHealthBlock)
        self.gridLayout_4.addWidget(self.pushButton_38, 7, 6, 1, 1)
        self.label_50 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_50.sizePolicy().hasHeightForWidth())
        self.label_50.setSizePolicy(sizePolicy)
        self.label_50.setMaximumSize(QtCore.QSize(16777215, 40))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(12)
        font.setItalic(True)
        self.label_50.setFont(font)
        self.label_50.setStyleSheet("QLabel{color : red;}")
        self.label_50.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_50.setObjectName("label_50")
        self.gridLayout_4.addWidget(self.label_50, 4, 6, 1, 4)
        self.label_47 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_47.sizePolicy().hasHeightForWidth())
        self.label_47.setSizePolicy(sizePolicy)
        self.label_47.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_47.setFont(font)
        self.label_47.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_47.setAlignment(QtCore.Qt.AlignCenter)
        self.label_47.setObjectName("label_47")
        self.gridLayout_4.addWidget(self.label_47, 8, 3, 1, 3)
        self.pushButton_39 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_39.sizePolicy().hasHeightForWidth())
        self.pushButton_39.setSizePolicy(sizePolicy)
        self.pushButton_39.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_39.setFont(font)
        self.pushButton_39.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_39.setObjectName("pushButton_39")
        self.pushButton_39.installEventFilter(TabWidget)
        self.pushButton_39.clicked.connect(self.onSelectSubDistrict)
        self.gridLayout_4.addWidget(self.pushButton_39, 9, 1, 1, 2)
        self.lineEdit_9 = QtWidgets.QLineEdit(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_9.sizePolicy().hasHeightForWidth())
        self.lineEdit_9.setSizePolicy(sizePolicy)
        self.lineEdit_9.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_9.setFont(font)
        self.lineEdit_9.setObjectName("lineEdit_9")
        self.gridLayout_4.addWidget(self.lineEdit_9, 1, 2, 1, 5)
        self.pushButton_40 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_40.sizePolicy().hasHeightForWidth())
        self.pushButton_40.setSizePolicy(sizePolicy)
        self.pushButton_40.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_40.setFont(font)
        self.pushButton_40.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_40.setObjectName("pushButton_40")
        self.pushButton_40.installEventFilter(TabWidget)
        self.pushButton_40.clicked.connect(self.onSelectDistrict)

        self.gridLayout_4.addWidget(self.pushButton_40, 8, 1, 1, 2)
        spacerItem2 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_4.addItem(spacerItem2, 3, 8, 1, 1)
        self.lineEdit_8 = QtWidgets.QLineEdit(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_8.sizePolicy().hasHeightForWidth())
        self.lineEdit_8.setSizePolicy(sizePolicy)
        self.lineEdit_8.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_8.setFont(font)
        self.lineEdit_8.setObjectName("lineEdit_8")
        self.gridLayout_4.addWidget(self.lineEdit_8, 3, 0, 1, 4)
        self.label_42 = QtWidgets.QLabel(self.tab1)
        self.label_42.setMaximumSize(QtCore.QSize(32, 32))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_42.setFont(font)
        self.label_42.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_42.setAlignment(QtCore.Qt.AlignCenter)
        self.label_42.setObjectName("label_42")
        self.gridLayout_4.addWidget(self.label_42, 3, 7, 1, 1)
        self.label_43 = QtWidgets.QLabel(self.tab1)
        self.label_43.setMaximumSize(QtCore.QSize(100, 32))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(10)
        font.setItalic(True)
        self.label_43.setFont(font)
        self.label_43.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_43.setAlignment(QtCore.Qt.AlignCenter)
        self.label_43.setObjectName("label_43")
        self.gridLayout_4.addWidget(self.label_43, 11, 9, 1, 1)
        spacerItem3 = QtWidgets.QSpacerItem(40, 20, QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        self.gridLayout_4.addItem(spacerItem3, 1, 8, 1, 1)
        self.pushButton_36 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_36.sizePolicy().hasHeightForWidth())
        self.pushButton_36.setSizePolicy(sizePolicy)
        self.pushButton_36.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_36.setFont(font)
        self.pushButton_36.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_36.setObjectName("pushButton_36")
        self.pushButton_36.installEventFilter(TabWidget)
        self.pushButton_36.clicked.connect(self.onSelectFacilityName)

        self.gridLayout_4.addWidget(self.pushButton_36, 11, 6, 1, 1)
        self.pushButton_43 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_43.sizePolicy().hasHeightForWidth())
        self.pushButton_43.setSizePolicy(sizePolicy)
        self.pushButton_43.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_43.setFont(font)
        self.pushButton_43.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_43.setObjectName("pushButton_43")
        self.pushButton_43.installEventFilter(TabWidget)
        self.pushButton_43.clicked.connect(self.onSelectBlock)
        self.gridLayout_4.addWidget(self.pushButton_43, 11, 1, 1, 2)
        self.pushButton_42 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_42.sizePolicy().hasHeightForWidth())
        self.pushButton_42.setSizePolicy(sizePolicy)
        self.pushButton_42.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_42.setFont(font)
        self.pushButton_42.setStyleSheet("background-color: #B00020;color: white;")
        self.pushButton_42.setObjectName("pushButton_42")
        self.pushButton_42.clicked.connect(self.reset)
        
        self.gridLayout_4.addWidget(self.pushButton_42, 9, 9, 1, 1)
        self.label_52 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_52.sizePolicy().hasHeightForWidth())
        self.label_52.setSizePolicy(sizePolicy)
        self.label_52.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_52.setFont(font)
        self.label_52.setStyleSheet("color: grey;background-color: rgb(255, 255, 255);")
        self.label_52.setAlignment(QtCore.Qt.AlignCenter)
        self.label_52.setObjectName("label_52")
        self.gridLayout_4.addWidget(self.label_52, 7, 0, 1, 1)
        self.pushButton_32 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_32.sizePolicy().hasHeightForWidth())
        self.pushButton_32.setSizePolicy(sizePolicy)
        self.pushButton_32.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_32.setFont(font)
        self.pushButton_32.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_32.setObjectName("pushButton_32")
        self.gridLayout_4.addWidget(self.pushButton_32, 8, 6, 1, 1)
        self.pushButton_32.installEventFilter(TabWidget)
        self.pushButton_32.clicked.connect(self.onSelectRuralUrban)

        self.label_40 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_40.sizePolicy().hasHeightForWidth())
        self.label_40.setSizePolicy(sizePolicy)
        self.label_40.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_40.setFont(font)
        self.label_40.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_40.setAlignment(QtCore.Qt.AlignCenter)
        self.label_40.setObjectName("label_40")
        self.gridLayout_4.addWidget(self.label_40, 11, 0, 1, 1)
        self.lineEdit_7 = QtWidgets.QLineEdit(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lineEdit_7.sizePolicy().hasHeightForWidth())
        self.lineEdit_7.setSizePolicy(sizePolicy)
        self.lineEdit_7.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_7.setFont(font)
        self.lineEdit_7.setObjectName("lineEdit_7")
        self.gridLayout_4.addWidget(self.lineEdit_7, 3, 4, 1, 2)
        self.label_41 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_41.sizePolicy().hasHeightForWidth())
        self.label_41.setSizePolicy(sizePolicy)
        self.label_41.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_41.setFont(font)
        self.label_41.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_41.setAlignment(QtCore.Qt.AlignCenter)
        self.label_41.setObjectName("label_41")
        self.gridLayout_4.addWidget(self.label_41, 9, 0, 1, 1)
        self.pushButton_35 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_35.sizePolicy().hasHeightForWidth())
        self.pushButton_35.setSizePolicy(sizePolicy)
        self.pushButton_35.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_35.setFont(font)
        self.pushButton_35.setStyleSheet("background-color: #003679;color: white;")
        self.pushButton_35.setObjectName("pushButton_35")
        self.pushButton_35.clicked.connect(self.VerifyFType)
        self.gridLayout_4.addWidget(self.pushButton_35, 3, 6, 1, 1)
        self.pushButton_34 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_34.sizePolicy().hasHeightForWidth())
        self.pushButton_34.setSizePolicy(sizePolicy)
        self.pushButton_34.setMaximumSize(QtCore.QSize(16777215, 55))
        font = QtGui.QFont()
        font.setPointSize(14)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_34.setFont(font)
        self.pushButton_34.setStyleSheet("QPushButton{background-color: #F47B1F;color: white;}\n"
"\n"
"")
        self.pushButton_34.setObjectName("pushButton_34")
        self.pushButton_34.clicked.connect(self.guidelines)
        self.gridLayout_4.addWidget(self.pushButton_34, 1, 9, 1, 1)
        self.label_39 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Expanding, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_39.sizePolicy().hasHeightForWidth())
        self.label_39.setSizePolicy(sizePolicy)
        self.label_39.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.label_39.setFont(font)
        self.label_39.setStyleSheet("color: grey;\n"
"background-color: rgb(255, 255, 255);")
        self.label_39.setAlignment(QtCore.Qt.AlignCenter)
        self.label_39.setObjectName("label_39")
        self.gridLayout_4.addWidget(self.label_39, 8, 0, 1, 1)
        self.pushButton_33 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Expanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_33.sizePolicy().hasHeightForWidth())
        self.pushButton_33.setSizePolicy(sizePolicy)
        self.pushButton_33.setMaximumSize(QtCore.QSize(16777215, 50))
        font = QtGui.QFont()
        font.setPointSize(9)
        self.pushButton_33.setFont(font)
        self.pushButton_33.setStyleSheet("background-color: rgb(222, 222, 222);")
        self.pushButton_33.setObjectName("pushButton_33")
        self.gridLayout_4.addWidget(self.pushButton_33, 9, 6, 1, 1)
        self.pushButton_33.installEventFilter(TabWidget)
        self.pushButton_33.clicked.connect(self.onSelectOwnership)

        self.label_57 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_57.sizePolicy().hasHeightForWidth())
        self.label_57.setSizePolicy(sizePolicy)
        self.label_57.setMaximumSize(QtCore.QSize(100, 32))
        font = QtGui.QFont()
        font.setFamily("MS Sans Serif")
        font.setPointSize(10)
        font.setItalic(True)
        self.label_57.setFont(font)
        self.label_57.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_57.setAlignment(QtCore.Qt.AlignCenter)
        self.label_57.setObjectName("label_57")
        self.gridLayout_4.addWidget(self.label_57, 5, 4, 2, 2)
        self.label_48 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_48.sizePolicy().hasHeightForWidth())
        self.label_48.setSizePolicy(sizePolicy)
        self.label_48.setMaximumSize(QtCore.QSize(16777215, 16777215))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.label_48.setFont(font)
        self.label_48.setStyleSheet("QLabel{background-color: white;color: grey;}")
        self.label_48.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_48.setObjectName("label_48")
        self.gridLayout_4.addWidget(self.label_48, 5, 0, 2, 4)
        self.label_44 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_44.sizePolicy().hasHeightForWidth())
        self.label_44.setSizePolicy(sizePolicy)
        self.label_44.setMaximumSize(QtCore.QSize(32, 36))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_44.setFont(font)
        self.label_44.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_44.setAlignment(QtCore.Qt.AlignCenter)
        self.label_44.setObjectName("label_44")
        self.gridLayout_4.addWidget(self.label_44, 9, 8, 1, 1)
        self.pushButton_41 = QtWidgets.QPushButton(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.MinimumExpanding, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_41.sizePolicy().hasHeightForWidth())
        self.pushButton_41.setSizePolicy(sizePolicy)
        self.pushButton_41.setMaximumSize(QtCore.QSize(16777215, 65))
        font = QtGui.QFont()
        font.setPointSize(16)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_41.setFont(font)
        self.pushButton_41.setStyleSheet("background-color: #73C067;color: white;")
        self.pushButton_41.setObjectName("pushButton_41")
        self.pushButton_41.clicked.connect(self.export)

        self.gridLayout_4.addWidget(self.pushButton_41, 8, 9, 1, 1)
        self.label_49 = QtWidgets.QLabel(self.tab1)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.MinimumExpanding)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_49.sizePolicy().hasHeightForWidth())
        self.label_49.setSizePolicy(sizePolicy)
        self.label_49.setMaximumSize(QtCore.QSize(32, 32))
        font = QtGui.QFont()
        font.setBold(True)
        font.setWeight(75)
        self.label_49.setFont(font)
        self.label_49.setStyleSheet("QLabel{background-color: rgb(255, 189, 167);}")
        self.label_49.setAlignment(QtCore.Qt.AlignCenter)
        self.label_49.setObjectName("label_49")
        self.gridLayout_4.addWidget(self.label_49, 8, 8, 1, 1)
        self.gridLayout_4.setColumnStretch(0, 1)
        self.gridLayout_4.setColumnStretch(2, 1)
        self.gridLayout_4.setColumnStretch(5, 1)
        self.gridLayout_4.setColumnStretch(6, 1)
        self.gridLayout_4.setColumnStretch(9, 1)
        self.gridLayout_4.setRowStretch(0, 2)
        self.gridLayout_4.setRowStretch(1, 1)
        self.gridLayout_4.setRowStretch(3, 1)
        self.gridLayout_4.setRowStretch(6, 1)
        self.gridLayout_4.setRowStretch(7, 1)
        self.gridLayout_4.setRowStretch(8, 1)
        self.gridLayout_4.setRowStretch(9, 1)
        self.gridLayout_4.setRowStretch(11, 1)
        self.gridLayout_5.addLayout(self.gridLayout_4, 0, 0, 1, 1)
        TabWidget.addTab(self.tab1, "")

        self.retranslateUi(TabWidget)
        TabWidget.setCurrentIndex(0)
        QtCore.QMetaObject.connectSlotsByName(TabWidget)

    def retranslateUi(self, TabWidget):
        _translate = QtCore.QCoreApplication.translate
        TabWidget.setWindowTitle(_translate("TabWidget", "TabWidget"))
        self.label_7.setText(_translate("TabWidget", "District"))
        self.pushButton_11.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_12.setText(_translate("TabWidget", "-- All Selected --"))
        self.label_11.setText(_translate("TabWidget", "Block"))
        self.label_9.setText(_translate("TabWidget", "Sub-District"))
        self.lineEdit_3.setPlaceholderText(_translate("TabWidget", " Month, Year"))
        self.pushButton_2.setText(_translate("TabWidget", "Guidelines"))
        self.pushButton_5.setText(_translate("TabWidget", "Validate"))
        self.pushButton_13.setText(_translate("TabWidget", "-- All Selected --"))
        self.lineEdit_2.setPlaceholderText(_translate("TabWidget", "  Facility Type selected will display here ..."))
        self.label_13.setText(_translate("TabWidget", "2**"))
        self.label_18.setText(_translate("TabWidget", "** Mandatory"))
        self.pushButton.setText(_translate("TabWidget", "Upload"))
        self.label_2.setText(_translate("TabWidget", "1**"))
        self.label_6.setText(_translate("TabWidget", "Health Block"))
        self.pushButton_10.setText(_translate("TabWidget", "-- All Selected --"))
        self.label_8.setText(_translate("TabWidget", "Rural / Urban"))
        self.lineEdit.setPlaceholderText(_translate("TabWidget", "  Your uploaded file name will display here ..."))
        self.label_4.setText(_translate("TabWidget", "Select Filters"))
        self.label_19.setText(_translate("TabWidget", "* Press Validate button to perform validation check on your data"))
        self.pushButton_8.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_7.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_15.setText(_translate("TabWidget", "Reset"))
        self.label_5.setText(_translate("TabWidget", "State"))
        self.pushButton_9.setText(_translate("TabWidget", "-- All Selected --"))
        self.pushButton_6.setText(_translate("TabWidget", "-- All Selected --"))
        self.label_10.setText(_translate("TabWidget", "Ownership"))
        self.label_12.setText(_translate("TabWidget", "Facility Name"))
        self.label.setText(_translate("TabWidget", "Data Validation Tool"))
        self.label_3.setText(_translate("TabWidget", "* Upload data in .xls / .xlsx format for one month and one facility type only."))
        self.pushButton_4.setText(_translate("TabWidget", "User Manual"))
        self.label_14.setText(_translate("TabWidget", "3 (optional)"))
        self.label_16.setText(_translate("TabWidget", "5**"))
        self.pushButton_14.setText(_translate("TabWidget", "Download"))
        self.label_15.setText(_translate("TabWidget", "4**"))
        TabWidget.setTabText(TabWidget.indexOf(self.tab), _translate("TabWidget", "English Version"))
        self.pushButton_44.setText(_translate("TabWidget", "-- ????????? ??????????????? --"))
        self.label_53.setText(_translate("TabWidget", "??????????????????"))
        self.label_54.setText(_translate("TabWidget", "??????????????????????????? ?????????????????? ?????? ?????????"))
        self.label_55.setText(_translate("TabWidget", "???????????? ???????????????????????? ?????????"))
        self.label_56.setText(_translate("TabWidget", "* ???????????? ?????? ?????????????????? ?????? ??????????????????????????? ?????????????????? ?????? ???????????? ?????? ???????????????  ?????? ????????? .xls / .xlsx ????????????????????? ????????? ???????????? ??????????????? ???????????????"))
        self.pushButton_45.setText(_translate("TabWidget", "????????????????????????"))
        self.pushButton_37.setText(_translate("TabWidget", "??????????????? ????????????"))
        self.label_45.setText(_translate("TabWidget", "1**"))
        self.label_46.setText(_translate("TabWidget", "??????????????? ???????????????"))
        self.pushButton_38.setText(_translate("TabWidget", "-- ????????? ??????????????? --"))
        self.label_50.setText(_translate("TabWidget", "* ???????????? ???????????? ?????? ????????????????????? ???????????? ???????????? ?????? ????????? ??????????????? ????????? ???????????????"))
        self.label_47.setText(_translate("TabWidget", "???????????? / ?????????????????????"))
        self.pushButton_39.setText(_translate("TabWidget", "-- ????????? ??????????????? --"))
        self.lineEdit_9.setPlaceholderText(_translate("TabWidget", "???????????? ??????????????? ?????? ?????? ??????????????? ?????? ????????? ???????????? ??????????????????????????? ????????????..."))
        self.pushButton_40.setText(_translate("TabWidget", "-- ????????? ??????????????? --"))
        self.lineEdit_8.setPlaceholderText(_translate("TabWidget", "??????????????? ??????????????????????????? ?????????????????? ?????????????????? ???????????? ??????????????????????????? ????????????..."))
        self.label_42.setText(_translate("TabWidget", "2**"))
        self.label_43.setText(_translate("TabWidget", "** ????????????????????????"))
        self.pushButton_36.setText(_translate("TabWidget", "-- ????????? ??????????????? --"))
        self.pushButton_43.setText(_translate("TabWidget", "-- ????????? ??????????????? --"))
        self.pushButton_42.setText(_translate("TabWidget", "???????????? ??????????????? ????????????"))
        self.label_52.setText(_translate("TabWidget", "???????????????"))
        self.pushButton_32.setText(_translate("TabWidget", "-- ????????? ??????????????? --"))
        self.label_40.setText(_translate("TabWidget", "???????????????"))
        self.lineEdit_7.setPlaceholderText(_translate("TabWidget", "???????????????,  ????????????"))
        self.label_41.setText(_translate("TabWidget", "?????? - ????????????"))
        self.pushButton_35.setText(_translate("TabWidget", "???????????? ??????????????? ????????????"))
        self.pushButton_34.setText(_translate("TabWidget", "?????????????????????????????????"))
        self.label_39.setText(_translate("TabWidget", "????????????"))
        self.pushButton_33.setText(_translate("TabWidget", "-- ????????? ??????????????? --"))
        self.label_57.setText(_translate("TabWidget", "3 (????????????????????????)"))
        self.label_48.setText(_translate("TabWidget", "????????????????????? ?????? ????????? ????????????"))
        self.label_44.setText(_translate("TabWidget", "5**"))
        self.pushButton_41.setText(_translate("TabWidget", "????????????????????? ????????????"))
        self.label_49.setText(_translate("TabWidget", "4**"))
        TabWidget.setTabText(TabWidget.indexOf(self.tab1), _translate("TabWidget", "??????????????? ?????????????????????"))


    ''' Upload Function '''
    # =====================

    def upload(self):
        global df_, res_dict, len_df_SummReport

        # Validation for uploaded valid excel file
        try:
            # Upload file by opening filedialog
            fileName,_ = QFileDialog.getOpenFileName(TabWidget, "Open Excel", (QtCore.QDir.homePath()), "Excel (*.xls *.xlsx)")
            
            # Read uploaded excel file
            df_ = pd.read_excel(fileName)

            # ## trim white spaces from column names
            # # iterating over the columns
            # for i in df_.columns:
            #     # applying strip function on column
            #     df_[i] = df_[i].str.strip()

            # Converted again to csv file
            df_.to_csv("FileName.csv")

            # Read converted csv file
            df_ = pd.read_csv("FileName.csv", skipinitialspace=True)

        except:
            msg = QMessageBox()
            msg.setWindowTitle("Uploaded File Error Message / ??????????????? ?????? ?????? ??????????????? ?????????????????? ???????????????")
            msg.setIcon(QMessageBox.Critical)
            msg.setText(
                "The file which you have uploaded is not in the valid format of excel, Please upload valid excel file \n\n ???????????? ?????????????????? ??????????????? ?????? ?????? ??????????????? ?????????????????? ?????? ??????????????? ????????????????????? ????????? ???????????? ??????, ??????????????? ??????????????? ?????????????????? ??????????????? ??????????????? ????????????")
            msg.exec()

            try:
                # Upload file by opening filedialog
                fileName, _ = QFileDialog.getOpenFileName(
                    TabWidget, "Open Excel", (QtCore.QDir.homePath()), "Excel (*.xls *.xlsx)")

                # Read uploaded excel file
                df_ = pd.read_excel(fileName)

                # Converted again to csv file
                df_.to_csv("FileName.csv")

                # Read converted csv file
                df_ = pd.read_csv("FileName.csv", skipinitialspace=True)
            except:
                msg = QMessageBox()
                msg.setIcon(QMessageBox.Critical)
                msg.setWindowTitle("Uploaded File Error Message / ??????????????? ?????? ?????? ??????????????? ?????????????????? ???????????????")
                msg.setText("Please upload valid excel file, Try again! \n\n ??????????????? ??????????????? ?????????????????? ??????????????? ??????????????? ????????????, ???????????? ?????????????????? ????????????!")
                msg.exec()

        

        self.popup.start_progress()

        # Filling upload your file in English Version
        self.lineEdit.setText(fileName)
        # Filling upload your file in Hindi Version
        self.lineEdit_9.setText(fileName)

        # Dropping last two rows
        df_.drop(df_.index[[-1, -2]], inplace=True)

        # Extracting string from 1st cell of dataframe
        str_to_extr_MonthYear = str(df_.iloc[0])

        # grab the first row for the header
        new_header = df_.iloc[0]

        # #take the data less the header row
        df_ = df_[0: -1]

        # set the header row as the df header
        df_.columns = new_header

        # Extracting Month , Year from string
        results = re.findall(
            r"[abceglnoprtuvyADFJMNOS|]{3}[\s-]\d{2,4}", str_to_extr_MonthYear)

        # # Splitting Month and Year
        # MYList = results[0].split('-')

        # cnt_lst1 = [i for i in df_.iteritems() if i == 'M1 [Ante Natal Care (ANC)]']

        l = df_.columns.values
        cnt = 0
        for i in l:
            if i != 'M1 [Ante Natal Care (ANC)]':
                cnt += 1
            else:
                break

        # Partial list of headers
        lst1 = df_.columns[:cnt].values

        # Picking row items after 18th row to merge with lst1
        lst2 = df_.iloc[1, cnt:].values.tolist()

        # Merging both lists
        lst3 = np.concatenate((lst1, lst2))

        # Assign lst3 as new column header
        df_.columns = lst3

        # Taking DataFrame from second row
        df_ = df_[3:]

        # Insering Month and Year to the orignal dataframe
        # df_.insert(1, 'Month', MYList[0])
        # df_.insert(2, 'Year', MYList[1])

        # Removing A column named as # coming from orignal data
        # df_ = df_.loc[:, df_.columns != '#']

        # Reindexing dataframe
        df_ = df_.reset_index(drop=True)

        df_ = df_.iloc[:, 1:]

        # convert the set to the list and fill inside comboBox to select facility type
        # df_.rename(columns={df_.filter(regex='^[F][a][c][i][l][i][t][y] [T][y][p][e]').columns[0]: 'FTYPE',},inplace=True)
        list_set = df_['Facility Type'].tolist()
        unique_list = set(list_set)

        # # Temporary column to verify modified checks
        temp_columns = ['col_' + str(index)
                        for index in range(1, len(df_.columns)+1)]
                        

        # Merging and converting temp_columns to orignal header to dictionary
        res_dict = {temp_columns[i]: df_.columns[i] for i in range(len(temp_columns))}


        # Picking the temporary column names and renaming column headers with it
        #df_.columns = [i for i in res_dict.keys()]

        # Orignal Header
        df_OrgHeaders = [i for i in res_dict.values()]

        try:
            # Filling Facility Name selected in English Version
            self.lineEdit_2.setText(["{0}".format(col) for col in unique_list][1])
            # Filling Month, Year in English Version
            self.lineEdit_3.setText(results[0])
            # Filling Facility Name selected in Hindiy Version
            self.lineEdit_8.setText(["{0}".format(col) for col in unique_list][1])
            # Filling Month, Year in Hindi Version
            self.lineEdit_7.setText(results[0])
        except:
            # Filling Facility Name selected in English Version
            self.lineEdit_2.setText(["{0}".format(col) for col in unique_list][0])
            # Filling Month, Year in English Version
            self.lineEdit_3.setText(results[0])
            # Filling Facility Name selected in Hindiy Version
            self.lineEdit_8.setText(["{0}".format(col) for col in unique_list][0])
            # Filling Month, Year in Hindi Version
            self.lineEdit_7.setText(results[0])
        finally:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("The file has been uploaded successfully. \n\n ??????????????? ????????????????????????????????? ??????????????? ?????? ?????? ?????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle("Error...!")
            # Display the message box
            self.msg.show()



        ## Removing null rows
        # df_ = df_.dropna(how='any',axis=1)

        len_df_SummReport = len(df_.columns)

        # Disabling upload Button
        self.pushButton.setDisabled(True)
        # self.pushButton_5.setDisabled(True)

        ############  Remove after USE #######################
        # df_.to_csv("NEW_FILENAME.csv")
        #######################################################

        self.popup.close()

        # Create the messagebox object
        self.msg = QMessageBox()
        # Set the information icon
        self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
        self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
        # Set the main message
        self.msg.setText("The file has been uploaded successfully. \n\n ??????????????? ????????????????????????????????? ??????????????? ?????? ?????? ?????? ?????????")
        # Set the title of the window
        self.msg.setWindowTitle(" ")
        # Display the message box
        self.msg.show()
        return df_


    # Upload file button functionality
    # ================================
    def loadFile(self, df_):
        return df_

    # Filtering Facility Type
    # =======================
    def VerifyFType(self):
        global df, FType
        self.popup.start_progress()

        FType = self.lineEdit_2.text()

        if (FType == 'Primary Health Centre'):
            # Signaling PHC_Validate function i.e function where validation checks are present
            df = PHC_Validate(self, df_)

        elif (FType == 'Health Sub Centre' ):
            # Signaling HSC_Validate function i.e function where validation checks are present
            df = HSC_Validate(self, df_)

        elif (FType == 'District Hospital'):
            df = DH_Validate(self, df_)

        elif (FType == 'Sub District Hospital'):
            df = SDH_Validate(self, df_)

        elif (FType == 'Community Health Centre'):
            df = CHC_Validate(self, df_)

        else:
            raise Exception('Facility Type Name is not matching')

        # self.pushButton_5.setEnabled(False)
        # self.pushButton_2.setEnabled(False)
        self.popup.close()

    '''
    # Filter to decide which filter button user clicked
    # =================================================
    '''
    def eventFilter(self, target, event):
        if target == self.pushButton_6 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_6.clicked.connect(self.onSelectState)
            return True

        elif target == self.pushButton_7 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_7.clicked.connect(self.onSelectDistrict)
            return True

        elif target == self.pushButton_13 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_13.clicked.connect(self.onSelectFacilityName)
            return True

        elif target == self.pushButton_11 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_11.clicked.connect(self.onSelectRuralUrban)
            return True

        elif target == self.pushButton_12 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_12.clicked.connect(self.onSelectOwnership)
            return True

        elif target == self.pushButton_8 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_8.clicked.connect(self.onSelectSubDistrict)
            return True

        elif target == self.pushButton_9 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_9.clicked.connect(self.onSelectBlock)
            return True

        elif target == self.pushButton_10 and event.type() == QtCore.QEvent.MouseButtonPress:
            self.pushButton_10.clicked.connect(self.onSelectHealthBlock)
            return True

        return False


    ################################################################################
    # Filter State Functionality
    
    def onSelectState(self, index):
        try:
            self.keywords = dict([(i, []) for i in range(df.shape[0])])
            print(self.keywords)
            self.menu = ScrollableMenu(maxItemCount=2000)
            self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
            font = self.menu.font()
            font.setPointSize(12)
            font.setBold(True)
            font.setWeight(75)
            self.menu.setFont(font)

            ## df.rename(columns={df.filter(regex='^[S][t][a][t][e]|^ +[S][t][a][t][e]/i').columns[0]: 'col_3',},inplace=True)
            index = df.columns.get_loc('col_3')
            self.col = index

            data_unique = []

            self.checkBoxs = []

            # Selectall added into Dropdown
            checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
            checkBox.setStyleSheet("color: red; spacing: 5px; font-size:20px;")
            # All the checkboxes are enabled to check
            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(checkBox)
            self.menu.addAction(checkableAction)
            checkBox.setChecked(True)
            checkBox.stateChanged.connect(self.slotSelectState)

            # list storing state data
            df['col_3'].fillna('Blank', inplace = True)
            list_set = df['col_3'].tolist()

            item = list_set
            item = sorted(list_set, key=str.upper)

            # looping to fill checkboxes, initially all checkboxes will be checked
            for i in range(len(item)):
                if item[i] not in data_unique:
                    data_unique.append(item[i])
                    checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                    checkBox.setChecked(True)
                    checkBox.setEnabled(False)
                    checkableAction = QtWidgets.QWidgetAction(self.menu)
                    checkableAction.setDefaultWidget(checkBox)
                    self.menu.addAction(checkableAction)
                    self.checkBoxs.append(checkBox)

            # Ok, cancel button
            btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                                QtCore.Qt.Vertical, self.menu)

            # ok selected
            btn.accepted.connect(self.menuCloseState)
            # rejected , nothing selected
            btn.rejected.connect(self.menu.close)

            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(btn)
            self.menu.addAction(checkableAction)
            self.pushButton_6.setMenu(self.menu)
            self.pushButton_44.setMenu(self.menu)

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("State Column does not exist \n\n ??????????????? ???????????? ??????????????? ???????????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass
        

    # method to check -> uncheck and vice versa
    def slotSelectState(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

            ## To enable and disable checkboxes on click of select/ deselect all
            if state == 0:
                checkbox.setEnabled(True)
            elif state == 2:
                checkbox.setEnabled(False)

    # after ok selected
    def menuCloseState(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        
        self.filterdataState()
        self.pushButton_6.setText('Selected')
        self.pushButton_6.setEnabled(False)
        self.pushButton_44.setText('???????????????')
        self.pushButton_44.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataState(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_3'].shape[0])])

        j = 0
        for j in range(df['col_3'].shape[0]):
            item = df['col_3'].tolist()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter District Functionality

    def onSelectDistrict(self, index):
        try:
            self.keywords = dict([(i, []) for i in range(df.shape[0])])
            print(self.keywords)
            self.menu = ScrollableMenu(maxItemCount=2000)
            self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
            font = self.menu.font()
            font.setPointSize(12)
            font.setBold(True)
            font.setWeight(75)
            self.menu.setFont(font)

            index = df.columns.get_loc('col_5')
            self.col = index

            data_unique = []

            self.checkBoxs = []

            # Selectall/deselectall added into Dropdown
            checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
            checkBox.setStyleSheet("color: red; spacing: 5px; font-size:20px;")
            # All the checkboxes are enabled to check
            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(checkBox)
            self.menu.addAction(checkableAction)
            checkBox.setChecked(True)
            checkBox.stateChanged.connect(self.slotSelectDistrict)

            # list storing state data
            df['col_5'].fillna('Blank', inplace = True)
            list_set = df['col_5'].to_list()

            item = list_set
            item = sorted(list_set, key=str.upper)

            # looping to fill checkboxes, initially all checkboxes will be checked
            for i in range(len(item)):
                if item[i] not in data_unique:
                    data_unique.append(item[i])
                    checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                    checkBox.setChecked(True)
                    checkBox.setEnabled(False)
                    checkableAction = QtWidgets.QWidgetAction(self.menu)
                    checkableAction.setDefaultWidget(checkBox)
                    self.menu.addAction(checkableAction)
                    self.checkBoxs.append(checkBox)

            # Ok, cancel button
            btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                                QtCore.Qt.Vertical, self.menu)

            # ok selected
            btn.accepted.connect(self.menuCloseDistrict)
            # rejected , nothing selected
            btn.rejected.connect(self.menu.close)

            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(btn)
            self.menu.addAction(checkableAction)
            self.pushButton_7.setMenu(self.menu)
            self.pushButton_40.setMenu(self.menu)

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("District Column does not exist \n\n ???????????? ???????????? ??????????????? ???????????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # method to check -> uncheck and vice versa
    def slotSelectDistrict(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

            ## To enable and disable checkboxes on click of select/ deselect all
            if state == 0:
                checkbox.setEnabled(True)
            elif state == 2:
                checkbox.setEnabled(False)

    # after ok selected
    def menuCloseDistrict(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataDistrict()
        self.pushButton_7.setText('Selected')
        self.pushButton_7.setEnabled(False)
        self.pushButton_40.setText('???????????????')
        self.pushButton_40.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataDistrict(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_5'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_5'].shape[0]):
            item = df['col_5'].tolist()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Sub District Functionality

    def onSelectSubDistrict(self, index):
        try:
            self.keywords = dict([(i, []) for i in range(df.shape[0])])
            print(self.keywords)
            self.menu = ScrollableMenu(maxItemCount=2000)
            self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
            font = self.menu.font()
            font.setPointSize(12)
            font.setBold(True)
            font.setWeight(75)
            self.menu.setFont(font)

            index = df.columns.get_loc('col_7')
            self.col = index

            data_unique = []

            self.checkBoxs = []

            # Selectall added into Dropdown
            checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
            checkBox.setStyleSheet("color: red; spacing: 5px; font-size:20px;")
            # All the checkboxes are enabled to check
            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(checkBox)
            self.menu.addAction(checkableAction)
            checkBox.setChecked(True)
            checkBox.stateChanged.connect(self.slotSelectSubDistrict)

            # list storing state data
            df['col_7'].fillna('Blank', inplace = True)
            list_set = df['col_7'].to_list()

            item = list_set
            item = sorted(list_set, key=str.upper)

            # looping to fill checkboxes, initially all checkboxes will be checked
            for i in range(len(item)):
                if item[i] not in data_unique:
                    data_unique.append(item[i])
                    checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                    checkBox.setChecked(True)
                    checkBox.setEnabled(False)
                    checkableAction = QtWidgets.QWidgetAction(self.menu)
                    checkableAction.setDefaultWidget(checkBox)
                    self.menu.addAction(checkableAction)
                    self.checkBoxs.append(checkBox)

            # Ok, cancel button
            btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                                QtCore.Qt.Vertical, self.menu)

            # ok selected
            btn.accepted.connect(self.menuCloseSubDistrict)
            # rejected , nothing selected
            btn.rejected.connect(self.menu.close)

            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(btn)
            self.menu.addAction(checkableAction)
            self.pushButton_8.setMenu(self.menu)
            self.pushButton_39.setMenu(self.menu)

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Sub District Column does not exist \n\n ?????? ???????????? ???????????? ??????????????? ???????????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # method to check -> uncheck and vice versa
    def slotSelectSubDistrict(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

            ## To enable and disable checkboxes on click of select/ deselect all
            if state == 0:
                checkbox.setEnabled(True)
            elif state == 2:
                checkbox.setEnabled(False)

    # after ok selected
    def menuCloseSubDistrict(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataSubDistrict()
        self.pushButton_8.setText('Selected')
        self.pushButton_8.setEnabled(False)
        self.pushButton_39.setText('???????????????')
        self.pushButton_39.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataSubDistrict(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_7'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_7'].shape[0]):
            item = df['col_7'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Block Functionality

    def onSelectBlock(self, index):
        try:
            self.keywords = dict([(i, []) for i in range(df.shape[0])])
            print(self.keywords)
            self.menu = ScrollableMenu(maxItemCount=2000)
            self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
            font = self.menu.font()
            font.setPointSize(12)
            font.setBold(True)
            font.setWeight(75)
            self.menu.setFont(font)

            index = df.columns.get_loc('col_8')
            self.col = index

            data_unique = []

            self.checkBoxs = []

            # Selectall added into Dropdown
            checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
            checkBox.setStyleSheet("color: red; spacing: 5px; font-size:20px;")
            # All the checkboxes are enabled to check
            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(checkBox)
            self.menu.addAction(checkableAction)
            checkBox.setChecked(True)
            checkBox.stateChanged.connect(self.slotSelectBlock)

            # list storing state data
            df['col_8'].fillna('Blank', inplace = True)
            list_set = df['col_8'].to_list()

            item = list_set
            item = sorted(list_set, key=str.upper)

            # looping to fill checkboxes, initially all checkboxes will be checked
            for i in range(len(item)):
                if item[i] not in data_unique:
                    data_unique.append(item[i])
                    checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                    checkBox.setChecked(True)
                    checkBox.setEnabled(False)
                    checkableAction = QtWidgets.QWidgetAction(self.menu)
                    checkableAction.setDefaultWidget(checkBox)
                    self.menu.addAction(checkableAction)
                    self.checkBoxs.append(checkBox)

            # Ok, cancel button
            btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                                QtCore.Qt.Vertical, self.menu)

            # ok selected
            btn.accepted.connect(self.menuCloseBlock)
            # rejected , nothing selected
            btn.rejected.connect(self.menu.close)

            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(btn)
            self.menu.addAction(checkableAction)
            self.pushButton_9.setMenu(self.menu)
            self.pushButton_43.setMenu(self.menu)

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Block Column does not exist \n\n ??????????????? ???????????? ??????????????? ???????????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # method to check -> uncheck and vice versa
    def slotSelectBlock(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

            ## To enable and disable checkboxes on click of select/ deselect all
            if state == 0:
                checkbox.setEnabled(True)
            elif state == 2:
                checkbox.setEnabled(False)

    # after ok selected
    def menuCloseBlock(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataBlock()
        self.pushButton_9.setText('Selected')
        self.pushButton_9.setEnabled(False)
        self.pushButton_43.setText('???????????????')
        self.pushButton_43.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataBlock(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_8'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_8'].shape[0]):
            item = df['col_8'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Health Block Functionality

    def onSelectHealthBlock(self, index):
        try : 
            self.keywords = dict([(i, []) for i in range(df.shape[0])])
            print(self.keywords)
            self.menu = ScrollableMenu(maxItemCount=2000)
            self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 200 }')
            font = self.menu.font()
            font.setPointSize(12)
            font.setBold(True)
            font.setWeight(75)
            self.menu.setFont(font)

            df.rename(columns={df.filter(regex='^[H][e][a][l][t][h] +[B][l][o][c][k]|^ +[H][e][a][l][t][h] +[B][l][o][c][k]|^[H][e][a][l][t][h] +[B][l][o][c][k]/i').columns[0]: 'col_9',},inplace=True)
            index = df.columns.get_loc('col_9')
            self.col = index

            data_unique = []

            self.checkBoxs = []

            # Selectall added into Dropdown
            checkBox = QtWidgets.QCheckBox("Select all / Deselect all", self.menu)
            checkBox.setStyleSheet("color: red; spacing: 5x; font-size:20px;")
            # All the checkboxes are enabled to check
            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(checkBox)
            self.menu.addAction(checkableAction)
            checkBox.setChecked(True)
            checkBox.stateChanged.connect(self.slotSelectHealthBlock)

            # list storing state data
            df['col_9'].fillna('Blank', inplace = True)
            list_set = df['col_9'].to_list()

            item = list_set
            item = sorted(list_set, key=str.upper)

            # looping to fill checkboxes, initially all checkboxes will be checked
            for i in range(len(item)):
                if item[i] not in data_unique:
                    data_unique.append(item[i])
                    checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                    checkBox.setChecked(True)
                    checkBox.setEnabled(False)
                    checkableAction = QtWidgets.QWidgetAction(self.menu)
                    checkableAction.setDefaultWidget(checkBox)
                    self.menu.addAction(checkableAction)
                    self.checkBoxs.append(checkBox)

            # Ok, cancel button
            btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                        QtCore.Qt.Vertical, self.menu)

            # ok selected
            btn.accepted.connect(self.menuCloseHealthBlock)
            # rejected , nothing selected
            btn.rejected.connect(self.menu.close)

            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(btn)
            self.menu.addAction(checkableAction)
            self.pushButton_10.setMenu(self.menu)
            self.pushButton_38.setMenu(self.menu)

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Health Block Column does not exist \n\n ??????????????? ??????????????? ???????????? ??????????????? ???????????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # method to check -> uncheck and vice versa
    def slotSelectHealthBlock(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

            ## To enable and disable checkboxes on click of select/ deselect all
            if state == 0:
                checkbox.setEnabled(True)
            elif state == 2:
                checkbox.setEnabled(False)

    # after ok selected
    def menuCloseHealthBlock(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataHealthBlock()
        self.pushButton_10.setText('Selected')
        self.pushButton_10.setEnabled(False)
        self.pushButton_38.setText('???????????????')
        self.pushButton_38.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataHealthBlock(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_9'].shape[0])])

        j = 0
        for j in range(df['col_9'].shape[0]):
            item = df['col_9'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Facility Name

    # Filter FacilityName Functionality

    def onSelectFacilityName(self, index):
        try:
            self.keywords = dict([(i, []) for i in range(df.shape[0])])
            print(self.keywords)
            # self.menu = QtWidgets.QMenu(TabWidget)
            self.menu = ScrollableMenu(maxItemCount=2000)
            self.menu.setStyleSheet('QMenu { menu-scrollable:true; width: 400 }')
            font = self.menu.font()
            font.setPointSize(10)
            font.setBold(True)
            font.setWeight(75)
            self.menu.setFont(font)

            ## df.rename(columns={df.filter(regex='^[F][a][c][i][l][i][t][y] +[N][a][m][e]|^ +[F][a][c][i][l][i][t][y] +[N][a][m][e]|^[F][a][c][i][l][i][t][y] +[N][a][m][e]/i').columns[0]: 'col_14',},inplace=True)
            index = df.columns.get_loc('col_14')
            self.col = index

            data_unique = []

            self.checkBoxs = []

            # Selectall added into Dropdown
            checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
            checkBox.setStyleSheet("color: red; spacing: 10px; font-size:20px;")
            # All the checkboxes are enabled to check
            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(checkBox)
            self.menu.addAction(checkableAction)
            checkBox.setChecked(True)
            checkBox.stateChanged.connect(self.slotSelectFacilityName)

            # list storing Facility Name data
            df['col_14'].fillna('Blank', inplace = True)
            list_set = df['col_14'].tolist()

            item = sorted(list_set, key=str.upper)

            # looping to fill checkboxes, initially all checkboxes will be checked
            for i in range(len(item)):
                if item[i] not in data_unique:
                    data_unique.append(item[i])
                    checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                    checkBox.setChecked(True)
                    checkBox.setEnabled(False)
                    checkableAction = QtWidgets.QWidgetAction(self.menu)
                    checkableAction.setDefaultWidget(checkBox)
                    self.menu.addAction(checkableAction)
                    self.checkBoxs.append(checkBox)

            # Ok, cancel button
            btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                                QtCore.Qt.Vertical, self.menu)

            # ok selected
            btn.accepted.connect(self.menuCloseFacilityName)
            # rejected , nothing selected
            btn.rejected.connect(self.menu.close)

            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(btn)
            self.menu.addAction(checkableAction)

            ############# Always set Pushbutton ####################
            self.pushButton_13.setMenu(self.menu)
            self.pushButton_36.setMenu(self.menu)

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Facility Name Column does not exist \n\n ??????????????????????????? ?????????????????? ?????? ????????? ???????????? ??????????????? ???????????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # method to check -> uncheck and vice versa
    def slotSelectFacilityName(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

            ## To enable and disable checkboxes on click of select/ deselect all
            if state == 0:
                checkbox.setEnabled(True)
            elif state == 2:
                checkbox.setEnabled(False)

    # after ok selected
    def menuCloseFacilityName(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataFacilityName()
        self.pushButton_13.setText('Selected')
        self.pushButton_13.setEnabled(False)
        self.pushButton_36.setText('???????????????')
        self.pushButton_36.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataFacilityName(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_14'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_14'].shape[0]):
            item = df['col_14'].tolist()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select District')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    ################################################################################
    # Filter Rural/Urban

    def onSelectRuralUrban(self, index):
        try:
            self.keywords = dict([(i, []) for i in range(df.shape[0])])
            self.menu = ScrollableMenu(maxItemCount=2000)
            self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
            font = self.menu.font()
            font.setPointSize(12)
            font.setBold(True)
            font.setWeight(75)
            self.menu.setFont(font)

            #df.rename(columns={df.filter(regex='^[R][u][r][a][l][/][U][r][b][a][n]|^ +[R][u][r][a][l][/][U][r][b][a][n]|^[R][u][r][a][l] +[/] +[U][r][b][a][n]|^[R][u][r][a][l] +[/][U][r][b][a][n]|^[R][u][r][a][l][/] +[U][r][b][a][n]/i').columns[0]: 'col_18',},inplace=True)
            index = df.columns.get_loc('col_18')
            self.col = index

            data_unique = []

            self.checkBoxs = []

            # Selectall added into Dropdown
            checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
            checkBox.setStyleSheet("color: red; spacing: 5px; font-size:20px;")
            # All the checkboxes are enabled to check
            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(checkBox)
            self.menu.addAction(checkableAction)
            checkBox.setChecked(True)
            checkBox.stateChanged.connect(self.slotSelectRuralUrban)

            # list storing Facility Name data
            df['col_18'].fillna('Blank', inplace = True)
            list_set = df['col_18'].to_list()

            item = list_set
            item = sorted(list_set, key=str.upper)

            # looping to fill checkboxes, initially all checkboxes will be checked
            for i in range(len(item)):
                if item[i] not in data_unique:
                    data_unique.append(item[i])
                    checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                    checkBox.setChecked(True)
                    checkBox.setEnabled(False)
                    checkableAction = QtWidgets.QWidgetAction(self.menu)
                    checkableAction.setDefaultWidget(checkBox)
                    self.menu.addAction(checkableAction)
                    self.checkBoxs.append(checkBox)

            # Ok, cancel button
            btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                                QtCore.Qt.Vertical, self.menu)

            # ok selected
            btn.accepted.connect(self.menuCloseRuralUrban)
            # rejected , nothing selected
            btn.rejected.connect(self.menu.close)

            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(btn)
            self.menu.addAction(checkableAction)

            ############# Always set Pushbutton ####################
            self.pushButton_11.setMenu(self.menu)
            self.pushButton_32.setMenu(self.menu)

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Rural/Urban Column does not exist \n\n ????????????/????????????????????? ???????????? ??????????????? ???????????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # method to check -> uncheck and vice versa
    def slotSelectRuralUrban(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

            ## To enable and disable checkboxes on click of select/ deselect all
            if state == 0:
                checkbox.setEnabled(True)
            elif state == 2:
                checkbox.setEnabled(False)

    # after ok selected
    def menuCloseRuralUrban(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataRuralUrban()
        self.pushButton_11.setText('Selected')
        self.pushButton_11.setEnabled(False)
        self.pushButton_32.setText('???????????????')
        self.pushButton_32.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataRuralUrban(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_18'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_18'].shape[0]):
            item = df['col_18'].tolist()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]
        print(final_lst, 'this is final list of Select Month')

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    ################################################################################
    # Select Ownership

    # Select Ownership Filter

    def onSelectOwnership(self, index):
        try:
            self.keywords = dict([(i, []) for i in range(df.shape[0])])
            print(self.keywords)
            self.menu = ScrollableMenu(maxItemCount=2000)
            self.menu.setStyleSheet('QMenu { menu-scrollable: true; width: 300 }')
            font = self.menu.font()
            font.setPointSize(12)
            font.setBold(True)
            font.setWeight(75)
            self.menu.setFont(font)

            #df.rename(columns={df.filter(regex='^[O][w][n][e][r][s][h][i][p]|^ +[O][w][n][e][r][s][h][i][p]/i').columns[0]: 'col_19',},inplace=True)
            index = df.columns.get_loc('col_19')
            self.col = index

            data_unique = []

            self.checkBoxs = []

            # Selectall added into Dropdown
            checkBox = QtWidgets.QCheckBox("Select all/ Deselect all", self.menu)
            checkBox.setStyleSheet("color: red; spacing: 5px; font-size:20px;")
            # All the checkboxes are enabled to check
            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(checkBox)
            self.menu.addAction(checkableAction)
            checkBox.setChecked(True)
            checkBox.stateChanged.connect(self.slotSelectOwnership)

            # list storing ownership column data
            df['col_19'].fillna('Blank', inplace = True)
            list_set = df['col_19'].to_list()

            item = list_set
            item = sorted(list_set, key=str.upper)

            # looping to fill checkboxes, initially all checkboxes will be checked
            for i in range(len(item)):
                if item[i] not in data_unique:
                    data_unique.append(item[i])
                    checkBox = QtWidgets.QCheckBox(item[i], self.menu)
                    checkBox.setChecked(True)
                    checkBox.setEnabled(False)
                    checkableAction = QtWidgets.QWidgetAction(self.menu)
                    checkableAction.setDefaultWidget(checkBox)
                    self.menu.addAction(checkableAction)
                    self.checkBoxs.append(checkBox)

            # Ok, cancel button
            btn = QtWidgets.QDialogButtonBox(QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel,
                                                QtCore.Qt.Vertical, self.menu)

            # ok selected
            btn.accepted.connect(self.menuCloseOwnership)
            # rejected , nothing selected
            btn.rejected.connect(self.menu.close)

            checkableAction = QtWidgets.QWidgetAction(self.menu)
            checkableAction.setDefaultWidget(btn)
            self.menu.addAction(checkableAction)

            ############# Always set Pushbutton ####################
            self.pushButton_12.setMenu(self.menu)
            self.pushButton_33.setMenu(self.menu)

        except:
            # Create the messagebox object
            self.msg = QMessageBox()
            # Set the information icon
            self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
            self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
            # Set the main message
            self.msg.setText("Ownership Column does not exist \n\n ?????????????????? ???????????? ??????????????? ???????????? ?????????")
            # Set the title of the window
            self.msg.setWindowTitle(" ")
            # Display the message box
            self.msg.show()
        
        finally:
            pass

    # method to check -> uncheck and vice versa
    def slotSelectOwnership(self, state):
        for checkbox in self.checkBoxs:
            checkbox.setChecked(QtCore.Qt.Checked == state)

            ## To enable and disable checkboxes on click of select/ deselect all
            if state == 0:
                checkbox.setEnabled(True)
            elif state == 2:
                checkbox.setEnabled(False)

    # after ok selected
    def menuCloseOwnership(self):
        self.keywords[self.col] = []
        for element in self.checkBoxs:
            if element.isChecked():
                self.keywords[self.col].append(element.text())
        print(self.keywords[self.col])
        self.filterdataOwnership()
        self.pushButton_12.setText('Selected')
        self.pushButton_12.setEnabled(False)
        self.pushButton_33.setText('???????????????')
        self.pushButton_33.setEnabled(False)
        self.menu.close()

    # Filter data columnwise
    def filterdataOwnership(self):
        global df
        #keywords = dict([(i, []) for i in range(self.filterall.columnCount())])
        columnsShow = dict([(i, True) for i in range(df['col_19'].shape[0])])
        print(columnsShow)

        j = 0
        for j in range(df['col_19'].shape[0]):
            item = df['col_19'].to_list()

            # if self.keywords[self.col]:
            if item[j] not in self.keywords[self.col]:
                columnsShow[j] = False

        # for key, value in columnsShow.items():
        final_lst = [i for i in columnsShow.values()]

        # matching list of facility type with col of dataframe returned by onSelectDistrict fun
        df = df[final_lst]
        return df


    # To count summary of the Modified Checks
    # =======================================
    def summaryReport(self, df):
        global final_result_summ1, final_result_summ2, col_sum, dft_ARFacilityWise, dft_ARCheckWiseInc, FList1, dft_ARCheckWisePRE, FList2, dft_FacilityWisePRE, dataframeForSheet6, dataframeForSheet7, FList3, FList4, dataframeForSheet4, dataframeForSheet5
        FType = self.lineEdit_2.text()

        # For Health Sub Centre
        if FType == 'Health Sub Centre': 
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 200th
            val_Description = [
                                'Number of mothers provided full course of 180 IFA tablets after delivery <= Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) + Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+ Number of Institutional Deliveries conducted',
                                'Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)<=Total number of pregnant women registered for ANC', 
                                'Out of the new cases of PW with hypertension detected, cases managed at institution <=New cases of PW with hypertension detected',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC',
                                'Number of PW tested using POC test for Syphilis<=Total number of pregnant women registered for ANC',
                                    'Out of the above, Number of PW found sero positive for Syphilis<=Number of PW tested using POC test for Syphilis',
                                    'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                    'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                    'Number of newborns received 6 HBNC?? visits after Institutional Delivery<=Number of Institutional Deliveries conducted',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted',
                                        'Child immunisation - Vitamin K1 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                        'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                            'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                            'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                            'Women receiving 1st post partum checkup within 48 hours of home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                            'Number of cases of AEFI - Abscess<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of cases of AEFI - Death<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of cases of AEFI - Others<=Number of children immunized (6.1.1+6.1.2+6.1.3+6.1.4+6.1.5+6.1.6+6.1.7+6.1.8+6.1.13+6.1.14+6.1.15+6.1.16+6.1.17+6.1.18+6.1.19+6.1.20+6.1.21+6.2.1+6.2.2+6.2.3+6.3.1+6.3.2+6.3.3+6.4.1+6.4.2+6.4.3+6.4.5+6.4.6+6.5.1+6.5.2+6.5.3+6.5.4)',
                                                'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held',
                                                'Out of the total number of Hb tests done, Number having Hb < 7 mg <=Number of Hb tests conducted',
                                                'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                    'Live Birth - Male+Live Birth - Female+Still Birth>=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted',
                                                    'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                    'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= 9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8',
                                                    'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria' ]

        # For Primary Health Centre
        elif FType == 'Primary Health Centre':
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 305th
            val_Description = [
                                'Child immunisation - Vitamin K (Birth Dose) <= Live Birth - Male+Live Birth - Female',
                                'Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC', 
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC', 
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC', 
                                'Number of PW given Tablet Misoprostol during home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Number of newborns received 7 Home Based Newborn Care (HBNC) visits in case of Home delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                    'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC', 
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Live Birth - Male+Live Birth - Female+Still Birth>=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception <=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth??<=Live Birth - Male+Live Birth - Female',
                                        'Women receiving 1st post partum checkup within 48 hours of home delivery<=Number of Home Deliveries attended by Skill Birth Attendant (SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of mothers provided 360 Calcium tablets after delivery<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Home Deliveries attended by Skill Birth Attendant(SBA) (Doctor/Nurse/ANM) +Number of Home Deliveries attended by Non SBA (Trained Birth Attendant(TBA) /Relatives/etc.)+Number of Institutional Deliveries conducted (Including C-Sections)',
                                                'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                                'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                                'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                                'Child immunisation - OPV 0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunized - Female<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                    'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                    'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                    'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                    'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                        'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= 14.1.1+14.1.2+14.1.3+14.1.4+14.1.5+14.1.6+14.1.7+14.1.8',
                                                        'Number of Left Against Medical Advice (LAMA) cases<=Male Admissions +Female Admissions',
                                                        'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'number positive for HIV (Number confirmed positive at ICTCs)<=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Number of cases of AEFI - Death<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Number of cases of AEFI - Others<=Number of Children Immunized (9.1.1+9.1.2+9.1.3+9.1.4+9.1.5+9.1.6+9.1.7+9.1.8+9.1.13+9.1.14+9.1.15+9.1.16+9.1.17+9.1.18+9.1.19+9.1.20+9.1.21+9.2.1+9.2.2+9.2.3+9.3.1+9.3.2+9.3.3+9.4.1+9.4.2+9.4.3+9.4.5+9.4.6+9.5.1+9.5.2+9.5.3+9.5.4)',
                                                            'Out of the new cases of PW with hypertension detected, cases managed at institution <=New cases of PW with hypertension detected',
                                                            'Immunisation sessions held <=Immunisation sessions planned',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held', 
                                                                'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria', 
                                                                'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria',
                                                                'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                                'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                                'Inpatient - Malaria <=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient Deaths - Male<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Inpatient Deaths - Female<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                                            'Male HIV - Number Positive<=Male HIV - Number Tested',
                                                                            'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis',
                                                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female']


        # For Sub District Hospital
        elif FType == 'Community Health Centre':
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 305th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                    'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                    'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                        'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                        'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                            'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                            'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                            'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                            'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                            'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                'Emergency - Burn<= Patients registered at Emergency Department',
                                                'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                    'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                    'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                    'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                    'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                    'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                    'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                    'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                    'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                    'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                    'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                    'Immunisation sessions held <=Immunisation sessions planned ',
                                                        'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                        'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                        'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                        'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                        'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                        'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                        'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                        'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn ??? Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                        'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                        'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                        'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis'
                                                        ]

        # For Primary Health Centre
        elif FType == 'Sub District Hospital':
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 321th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                    'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                        'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                        'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                            'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                            'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                                'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Tests Positive for JE<=Tests Conducted for JE',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology)',
                                                    'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                    'Emergency - Burn<= Patients registered at Emergency Department',
                                                    'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                        'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                        'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                        'Emergency - CVA ( Cerebovascular Disease)<= Patients registered at Emergency Department',
                                                        'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Immunisation sessions held <=Immunisation sessions planned ',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                            'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                            'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                            'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                            'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn ??? Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                            'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis']
            
        # For District Hospital
        elif FType == 'District Hospital':
            df_SummReport = df.iloc[:, len_df_SummReport:]     ## Taking columns after 326th
            val_Description = [
                                'Out of the ANC registered, number registered with in 1st trimester(Within 12 weeks)<=Total number of pregnant women registered for ANC',
                                'Male HIV Number Positive <= Male HIV - Number Tested',
                                'Number of PW given 180 Iron Folic Acid (IFA) tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW given 360 Calcium tablets <=Total number of pregnant women registered for ANC ',
                                'Number of PW received 4 or more ANC check ups<=Total number of pregnant women registered for ANC ',
                                'Out of the new cases of PW with hypertension detected, cases managed at institution<=New cases of PW with hypertension detected ',
                                'Number of Eclampsia cases managed during delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                'No. of PW having severe anaemia (Hb<7) treated could be greater than No. of PW having severe anaemia (Hb<7)  tested cases',
                                'Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)<=Total number of pregnant women registered for ANC ',
                                    'Number of PW tested positive for GDM<=Number of PW tested for Blood Sugar using OGTT (Oral glucose tolerance test)',
                                    'Number of PW given insulin out of total tested positive for GDM<=Number of PW tested positive for GDM',
                                    'Number of Pregnant women tested for Syphilis<=Total number of pregnant women registered for ANC ',
                                    'Number of Pregnant women tested found sero positive for Syphilis<=Number of Pregnant women tested for Syphilis',
                                    'Number of Syphilis positive pregnant women treated for Syphilis<=Number of Pregnant women tested found sero positive for Syphilis',
                                    'Number of babies treated for congenital Syphilis<=Number of babies diagnosed with congenital Syphilis',
                                    'Out of total institutional deliveries number of women discharged within 48 hours of delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'Total C -Section deliveries performed<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                    'C-sections, performed at night (8 PM- 8 AM)<=Total C -Section deliveries performed',
                                    'Live Birth - Male + Live Birth - Female + Still Birth>=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of Pre term newborns ( < 37 weeks of pregnancy)<=Live Birth - Male+Live Birth - Female',
                                        'Post Abortion/ MTP Complications Identified<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Post Abortion/ MTP Complications Treated<=Post Abortion/ MTP Complications Identified',
                                        'Number of women provided with post abortion/ MTP contraception<=MTP up to 12 weeks of pregnancy+MTP more than 12 weeks of pregnancy+Abortion (spontaneous)',
                                        'Number of newborns weighed at birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of newborns having weight less than 2.5 kg<=Number of newborns weighed at birth',
                                        'Number of Newborns breast fed within 1 hour of birth<=Live Birth - Male+Live Birth - Female',
                                        'Number of Complicated pregnancies treated with Blood Transfusion<=Number of cases of pregnant women with Obstetric Complications attended (Antepartum haemorrhage (APH), Post-Partum Hemorrhage (PPH), Sepsis, Eclampsia and others) ',
                                        'Number of mothers provided full course of 180 IFA tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'Number of mothers provided 360 Calcium tablets after delivery<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                        'RTI/STI for which treatment initiated - Male<=New RTI/STI cases identified - Male',
                                            'RTI/STI for which treatment initiated -Female<=New RTI/STI cases identified - Female',
                                            'Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of Post Partum (within 48 hours of delivery) IUCD insertions<=Number of Institutional Deliveries conducted (Including C-Sections)',
                                            'Number of complications following IUCD Insertion<=Number of Interval IUCD Insertions (excluding PPIUCD and PAIUCD)+ Number of post partum (with in 48 hours of delivery) IUCD insertion +Number of post abortion (with 12 days of spontaneous or surgical abortions)  IUCD incertion',
                                            'Complications following male sterilization<=Number of Non Scalpel Vasectomy (NSV) / Conventional Vasectomy conducted',
                                            'Complications following female sterilization<=Number of Laparoscopic sterilizations (excluding post abortion) conducted + Number of Interval Mini-lap (other than post-partum and post abortion) sterilizations conducted + Number of Post Partum sterilizations (within 7 days of delivery by minilap or concurrent with cessarean section) conducted + Number of Post Abortion sterilizations (within 7 days of spontaneous or surgical abortion) conducted',
                                            'Child immunisation - Vitamin K1(Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - BCG<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - OPV-0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Child immunisation - Hepatitis-B0 (Birth Dose)<=Live Birth - Male+Live Birth - Female',
                                            'Children aged between 9 and 11 months fully immunized- Male+Children aged between 9 and 11 months fully immunize<=Child immunisation (9-11months) - Measles & Rubella (MR) 1st dose  & Child immunisation (9-11months) - Measles 1st dose',
                                                'Kala Azar Positive Cases<=Kala Azar (RDT) - Tests Conducted',
                                                'Tests Positive for JE<=Tests Conducted for JE',
                                                'Out of registered, Girls received clinical services<=Girls registered in AFHC',
                                                'Out of registered, Boys received clinical services<=Boys registered in AFHC',
                                                'Out of registered, Girls received counselling<=Girls registered in AFHC',
                                                'Out of registered, Boys received counselling<=Boys registered in AFHC',
                                                'Allopathic- Outpatient attendance+Ayush - Outpatient attendance >= Number of outpatients (Diabetes + Hypertension +  Stroke (Paralysis) + Acute Heart Diseases + Mental illness + Epilepsy + Ophthalmic Related + Dental + Oncology',
                                                'Number of Left Against Medical Advice (LAMA) cases<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Malaria<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Dengue<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                'Inpatient - Typhoid<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Asthma, Chronic Obstructive Pulmonary Disease (COPD), Respiratory infections<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Tuberculosis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Pyrexia of unknown origin (PUO)<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Diarrhea with dehydration<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Inpatient - Hepatitis<=Inpatient (Male)- Children<18yrs+Inpatient (Male)- Adults+Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                    'Emergency - Trauma ( accident, injury, poisoning etc)<= Patients registered at Emergency Department',
                                                    'Emergency - Burn<= Patients registered at Emergency Department',
                                                    'Emergency - Obstetrics complications<= Patients registered at Emergency Department',
                                                    'Emergency - Snake Bite<=Patients registered at Emergency Department',
                                                    'Emergency - Acute Caridiac Emergencies<= Patients registered at Emergency Department',
                                                    'Emergency - CVA ( Cerebovascular Disease)<= Patients registered at Emergency Department',
                                                        'Number of deaths occurring at Emergency Department<= Patients registered at Emergency Department',
                                                        'Number of children discharged with target weight gain from the NRCs<=Number of children admitted in NRC',
                                                        'Out of the total number of Hb tests done, Number having Hb < 7 mg<=Number of Hb tests conducted',
                                                        'Female Non ANC HIV - Number Positive<=Female Non ANC HIV - Number Tested',
                                                        'out of the above, Number screened positive<=Number of Pregnant Women screened for HIV',
                                                        'out of the above, Number screened positive, number confirmed with HIV infection at Integrated Counselling and Testing Centre (ICTC) <=out of the above, Number screened positive',
                                                        'Widal tests - Number Positive<=Widal tests - Number Tested',
                                                        'Number of cases of AEFI - Abscess<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Death<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Number of cases of AEFI - Others<=Number of Children Immunized (Vitamin K (Birth Dose) + BCG + DPT1 + DPT2 + DPT3 + Pentavalent 1 + Pentavalent 2 + Pentavalent 3 + Hepatitis-B0 (Birth Dose) + Hepatitis-B1 +  Hepatitis-B2 + Hepatitis-B3 + Inactivated Injectable Polio Vaccine 1 (IPV 1) + Inactivated Injectable Polio Vaccine 2 (IPV 2) + Rotavirus 1 + Rotavirus 2 + Rotavirus 3 + (9-11 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (9-11 months) - Measles 1st Dose + (9-11 months) - JE 1st dose + (after 12 months) - Measles & Rubella (MR)/ Measles containing vaccine(MCV) - 1st Dose + (after 12 months) - Measles 1st Dose + (after 12 months) - JE 1st dose + Measles & Rubella (MR)- 2nd Dose (16-24 months) + Measles 2nd dose (More than 16 months) + DPT 1st Booster + Measles, Mumps, Rubella (MMR) Vaccine + Number of children more than 16 months of age who received Japanese Encephalitis (JE) vaccine + Typhoid + Children more than 5 years received DPT5 (2nd Booster) + Children more than 10 years received TT10/ Td10 + Children more than 16 years received TT16/ Td16)',
                                                        'Immunisation sessions held <=Immunisation sessions planned ',
                                                            'Number of Immunisation sessions where ASHAs were present<=Immunisation sessions held ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Vivax test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (Microscopy Tests ) - Plasmodium Falciparum test positive<=Total Blood Smears Examined for Malaria ',
                                                            'Malaria (RDT) - Plasmodium Vivax test positive<=RDT conducted for Malaria',
                                                            'Malaria (RDT) - Plamodium Falciparum test positive<=RDT conducted for Malaria',
                                                            'Total number of blood units issued during the month>=Number of blood units issued (Excluding C-Section)',
                                                            'Inpatient Deaths - Male <=Inpatient (Male)- Children<18yrs+Inpatient (Male)',
                                                            'Inpatient Deaths - Female<=Inpatient (Female)- Children<18yrs+Inpatient (Female)- Adults',
                                                            'Number of deaths occurring at SNCU<=Special Newborn Care Unit (SNCU Admissions) - Inborn Male + Special Newborn Care Unit (SNCU Admissions) - Inborn Female + Outborn ??? Male + Outborn - Female + Number of newborns admitted in SNCU - referred by ASHA',
                                                            'Out of Operation major, Gynecology- Hysterectomy surgeries<=Operation major (General and spinal anaesthesia)',
                                                            'Major Surgeries excluding Obstetrics, Gynaecology and Opthalmology etc.<=Operation major (General and spinal anaesthesia)',
                                                            'Number of Male STI/RTI attendees found sero Positive for syphilis<=Number of Male STI/RTI attendees tested for syphilis',
                                                            'Number of Female (Non ANC) STI/RTI attendees found sero Positive for syphilis<=Number of Female (Non ANC)STI/RTI attendees tested for syphilis']

        

        '''
        ## First Summary Report
        ## ---------------------
        '''

        # count_Consistent = []
        count_Inconsistent = []
        # count_Blank = []
        count_ProbableRErr = []



        Columns = list(df_SummReport.columns.values.tolist())

        for col_name in Columns:

            c2 = df_SummReport[col_name].str.count("Inconsistent").sum()
            count_Inconsistent.append(c2)

            c4 = df_SummReport[col_name].str.count("Probable Reporting Error").sum()
            count_ProbableRErr.append(c4)

        print(len(val_Description), len(count_Inconsistent), len(count_ProbableRErr))   

        # To show facilities in a column
        colInterest = df['col_14'].tolist() # Give column names in addition

        ''' For Inconsistent '''
        inconsistent_list = []

        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]
            colComparison = colComparison.tolist()
            for j in range(0, len(colComparison)):
                primString = colComparison[j]
                
                pattern = re.compile("^I")
                if pattern.match(str(primString)):
                    temp.append(colInterest[j])
                else:
                    continue

            inconsistent_list.append(temp)           

        ''' For PRE '''
        PRE_list = []
        # TO show facility in the sheet5
        lg = len(df_SummReport.columns)

        for i in range(0, lg):
            temp = []
            colComparison = df_SummReport.iloc[:,i]
            colComparison = colComparison.tolist()
            for j in range(0, len(colComparison)):
                primString = colComparison[j]

                pattern = re.compile("^P")
                if pattern.match(str(primString)):
                    temp.append(colInterest[j])
                else:
                    continue

            PRE_list.append(temp)      


        final_result_summ1 = pd.DataFrame({"Conditions": df_SummReport.columns, 
                                             "Description": val_Description,
                                                "Facilities(Name) Showing Inconsistent": inconsistent_list,
                                                    "Inconsistent": count_Inconsistent,
                                                        "Facilities (Name) Showing Probable Reporting Error": PRE_list,
                                                            "Probable Reporting Error": count_ProbableRErr,
                                                            })
        final_result_summ1 = final_result_summ1.sort_values(by=['Inconsistent'], ascending=False)
        final_result_summ1 = final_result_summ1.reset_index(drop=True)
                                                 
        FList1 = final_result_summ1["Facilities(Name) Showing Inconsistent"].tolist()
        FList2 = final_result_summ1["Facilities (Name) Showing Probable Reporting Error"].tolist()

        dataframeForSheet4 = final_result_summ1[['Conditions', 'Description', 'Inconsistent', 'Facilities(Name) Showing Inconsistent']]
        dataframeForSheet5 = final_result_summ1[['Conditions', 'Description', 'Probable Reporting Error', 'Facilities (Name) Showing Probable Reporting Error']]

        # Total number of rows in the upoaded dataset
        count_rows = df_SummReport.shape[0]
        
            
        # Percentage for Validation Summary Sheet to show color codes
        final_result_summ1['PerIncSheet1'] = final_result_summ1['Inconsistent']/count_rows*100
        final_result_summ1['Inconsistent'].value_counts()
        final_result_summ1['PerPRErrSheet1'] = final_result_summ1['Probable Reporting Error']/count_rows*100

        len0 = len(final_result_summ1['PerIncSheet1'])
        len1 = len(final_result_summ1['PerPRErrSheet1'])

        # Deleting unnecessary columns
        # del final_result_summ1['PerIncSheet1']
        # del final_result_summ1['PerPRErrSheet1']
            
        def select_col_SumSheet(X):
            # COLORS
            c = ['background-color:  #EF5350',                  #   >=25% RED
                        'background-color: #FFAF00',            #   10 - 25% LIGHTER RED
                            'background-color: #C0C000',        #   5 - 10% MORE LIGHTER RED
                                'background-color: #00AF5F',    #   < 5% LIGHTEST RED
                                    ' ']

            mask_30 = (X['Inconsistent'] == 0)
            cnt30 = mask_30.values.sum()
            mask_29 = (X['Probable Reporting Error'] == 0)
            cnt29 = mask_29.values.sum()

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[4], X.index, columns=X.columns)
            df1.loc[mask_29, 'Probable Reporting Error'] = c[3]
            df1.loc[mask_30, 'Inconsistent'] = c[3]

            return df1

        # Remoiving inconsistent facility names from validation summary sheet
        final_result_summ1.drop(['Facilities(Name) Showing Inconsistent'], axis = 1, inplace=True)
        # Remoiving PRE facility names from validation summary sheet
        final_result_summ1.drop(['Facilities (Name) Showing Probable Reporting Error'], axis = 1, inplace=True)

        final_result_summ1 = final_result_summ1.style.apply(select_col_SumSheet, axis=None)


        '''
        ## Second Summary Report
        ## --------------------
        '''
        summ2_countInconsistent = []
        summ2_countProbableRErr = []
        All_Blank = []

        # Iterating over indices of each row and calculating number of Blanks for each Facility Name 
        for index in range(len(df_SummReport)):
            '''   For no. of Inconsistent   '''
            inconsistent = df_SummReport.iloc[index, :].str.count("Inconsistent").sum()
            summ2_countInconsistent.append(inconsistent)
            
            '''   For no. of Probable Reporting Errors   '''
            probableRErr = df_SummReport.iloc[index, :].str.count('Probable Reporting Error').sum()
            summ2_countProbableRErr.append(probableRErr)

            blank = df_SummReport.iloc[index, :].str.count("Blank").sum()
            if blank == len(df_SummReport.columns):
                All_Blank.append('Yes')
            else:
                All_Blank.append('No')

        

        #########################################################   
        #  Facility Specific Inconsistent (Sheet 6)     
        
        ''' For Inconsistent '''
        # **********************
        inc_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]

                pattern = re.compile("^I")
                if pattern.match(str(primString)):
                    txt = primString.replace('Inconsistent', " ")
                    temp.append(txt)

            inc_list.append(temp)

        ''' For PRE '''
        # *************
        pre_list = []

        lg = len(df_SummReport.columns)
        len_df = df_SummReport.shape[0]

        for i in range(0, len_df):
            temp = []

            colComparison = df_SummReport.iloc[i,:]
            for j in range(0, lg):
                primString = colComparison[j]

                pattern = re.compile("^P")
                if pattern.match(str(primString)):
                    txt = primString.replace('Probable Reporting Error', " ")
                    temp.append(txt)

            pre_list.append(temp)



        ###################################################
        
        final_result_summ2 = pd.DataFrame({ "State": df['col_3'].tolist(),
                                                "District": df['col_5'].tolist(),
                                                    "Sub-district": df['col_7'].tolist(),
                                                        "Block": df['col_8'].tolist(),
                                                            "Facility Name/ Block/ Sub-district/ District/ State": colInterest,
                                                                "Inconsistent": summ2_countInconsistent,
                                                                    "Probable Reporting Error": summ2_countProbableRErr,
                                                                        "All Blank": All_Blank,
                                                                            "Checks (Inconsistent)" : inc_list,
                                                                                "Checks (PRE)": pre_list
                                                                                    })

        # Sorting in alphabetical  order
        # final_result_summ2 = final_result_summ2.sort_values(by=['Facility Name'], ascending=True)
        final_result_summ2 = final_result_summ2.sort_values(by=['Inconsistent'], ascending=False)
        final_result_summ2 = final_result_summ2.reset_index(drop=True)

        FList3 = final_result_summ2["Checks (Inconsistent)"].tolist()
        FList4 = final_result_summ2["Checks (PRE)"].tolist()

        '''
        ## 5th and 6th Summary Report (Top 10 Validation Checks not performing good)
        ## ------------------------------------------------------------------------
        '''
        dataframeForSheet6 = final_result_summ2[['Facility Name/ Block/ Sub-district/ District/ State', 'Block', 'Sub-district', 'District', 'State', 'Inconsistent', 'Checks (Inconsistent)']]
        dataframeForSheet6['Facility Name/ Block/ Sub-district/ District/ State'] = dataframeForSheet6['Facility Name/ Block/ Sub-district/ District/ State'] + " / " + dataframeForSheet6['Block'] + " / " + dataframeForSheet6['Sub-district'] + " / " + dataframeForSheet6['District'] + " / " + dataframeForSheet6['State'] 

        # dataframeForSheet6.to_csv('TEST_SHEET_1.csv')
        dataframeForSheet7 = final_result_summ2[['Facility Name/ Block/ Sub-district/ District/ State', 'Block', 'Sub-district', 'District', 'State', 'Probable Reporting Error', 'Checks (PRE)']]
        dataframeForSheet7['Facility Name/ Block/ Sub-district/ District/ State'] = dataframeForSheet7['Facility Name/ Block/ Sub-district/ District/ State'] + " / " + dataframeForSheet7['Block'] + " / " + dataframeForSheet7['Sub-district'] + " / " + dataframeForSheet7['District'] + " / " + dataframeForSheet7['State'] 


        '''  To find percentage Facility Type Wise   '''
                                                    
        # For Health Sub Centre
        if FType == 'Health Sub Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/31 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 31 * 100
        
        # For Primary Health Centre
        elif FType == 'Primary Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/78 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 78 * 100

        # For Community Health Centre
        elif FType == 'Community Health Centre':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/83 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 83 * 100

        # For Sub District Hospital
        elif FType == 'Sub District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/85 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 85 * 100

        # For District Hospital
        elif FType == 'District Hospital':
            final_result_summ2['PercentageInc'] = final_result_summ2['Inconsistent']/85 * 100
            final_result_summ2['PercentagePRErr'] = final_result_summ2['Probable Reporting Error']/ 85 * 100
        
        ## Deleting unnecessary columns
        del final_result_summ2['Checks (Inconsistent)']
        del final_result_summ2['Checks (PRE)']
        del final_result_summ2['Sub-district']
        del final_result_summ2['Block']

        def select_col(X):
            global c
            # COLORS
            # ******
            c = ['background-color:  #EF5350',                  #>=50% RED
                    'background-color: #FFAF00',                #25-50% ORANGE
                        'background-color: #C0C000',            #10-25% YELLOW
                            'background-color: #00FF00',        #5-10% L GREEN
                                'background-color: #00AF5F',    #<5% GREEN
                                    ' ']
   
                
            mask_AllBlank = (X['All Blank'] == 'Yes')   
            mask_16 = (X['Inconsistent'] == 0)
            mask_15 = (X['Probable Reporting Error'] == 0)

            #DataFrame with same index and columns names as original filled empty strings
            df1 =  pd.DataFrame(c[5], X.index, columns=X.columns)
            df1.loc[mask_AllBlank, 'All Blank'] = c[0]
            df1.loc[mask_15, 'Probable Reporting Error'] = c[4]
            df1.loc[mask_16, 'Inconsistent'] = c[4]
            return df1

        final_result_summ2 = final_result_summ2.style.apply(select_col, axis=None)

        return final_result_summ1, final_result_summ2, dataframeForSheet4, dataframeForSheet5, dataframeForSheet6, dataframeForSheet7


    # EXPORT FILE
    # ===========
    def export(self):
        table_result1, table_result2, table_result3, table_result4, table_result5, table_result6  = self.summaryReport(df)

        '''
        ADDING THICK BORDERS
        '''
        from openpyxl.styles.borders import Border, Side
        thick_border = Border(left=Side(style='thick'), 
                            right=Side(style='thick'), 
                            top=Side(style='thick'), 
                            bottom=Side(style='thick'))

        # Rename orignal headers
        # df.rename(res_dict , axis=1, inplace=True)

        ## RENAMING
        df.rename(columns={"col_14": "Facility Name", "col_3": "State", "col_5": "District Name", "col_7": "Sub-District Name", "col_8": "Block Name", "col_18": "Rural/Urban", "col_19": "Ownership",}, inplace=True)

        new_list = [["Data Validation Check Tool:"]]
        table_result_content = pd.DataFrame(new_list)


        # Save file dialog
        filename = QFileDialog.getSaveFileName(TabWidget, "Save to Excel", "Summary_Sheet",
                                                "Excel Spreadsheet (*.xlsx);;"
                                                "All Files (*)")[0]

        self.popup.start_progress()

        del table_result5['Block']
        del table_result5['Sub-district']
        del table_result5['District']
        del table_result5['State']

        del table_result6['Block']
        del table_result6['Sub-district']
        del table_result6['District']
        del table_result6['State']
        
        # Taking transpose of data 
        table_result3 = table_result3.T
        table_result4 = table_result4.T
        table_result5 = table_result5.T
        table_result6 = table_result6.T

        try:
        # exporting to excel
            with pd.ExcelWriter(filename) as writer: 
                table_result_content.to_excel(writer, sheet_name='Description', engine='openpyxl')
                table_result2.to_excel(writer, sheet_name='Facility Level Summary', engine='openpyxl')
                table_result5.to_excel(writer, sheet_name='Checks giving Inconsistent', engine='openpyxl')
                table_result6.to_excel(writer, sheet_name='Checks giving PRE', engine='openpyxl')
                table_result1.to_excel(writer, sheet_name='Validation checkwise summary', engine='openpyxl')
                table_result3.to_excel(writer, sheet_name='Facility with Inconsistent', engine='openpyxl')
                table_result4.to_excel(writer, sheet_name='Facility with PRE', engine='openpyxl')
                df.to_excel(writer, sheet_name='Validated_Data', engine='openpyxl')

        except:
            msg = QMessageBox()
            msg.setWindowTitle("Saving File Error Message / ??????????????? ?????????????????? ?????????????????? ???????????????")
            msg.setIcon(QMessageBox.Critical)
            msg.setText(
                "The file is already opened , CLOSE IT ... / ??????????????? ???????????? ?????? ???????????? ?????? ???????????? ??????, ????????? ????????? ????????????...")
            msg.setText(
                "\n WINDOWS PERMISSION DENIED !  The file is already opened , CLOSE IT FIRST ... / ????????????????????? ?????????????????? ????????????????????????! ??????????????? ???????????? ?????? ???????????? ?????? ???????????? ??????, ????????? ???????????? ????????? ????????????...")
            msg.exec()
        

        # PALETTES
        workbook = load_workbook(filename)
        sheet_0 = workbook['Description']
        sheet = workbook['Facility Level Summary']
        sheet_1 = workbook['Validation checkwise summary']
        sheet_2 = workbook['Facility with Inconsistent']
        sheet_3 = workbook['Facility with PRE']
        sheet_4 = workbook['Checks giving Inconsistent']
        sheet_5 = workbook['Checks giving PRE']
        sheet_6 = workbook['Validated_Data'] 

        # Activating sheets 
        workbook.active = sheet
        workbook.active = sheet_1
        workbook.active = sheet_4

        sheet.sheet_view.showGridLines = False
        sheet_1.sheet_view.showGridLines = False

        ''' THE CODE FOR GENERATING HYPERLINKS '''
        # Function to generate sequences according to the excel sheet
        def excel_cols_link3():
            n1 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n1))
                n1 += 1

        # Function to generate sequences according to the excel sheet
        def excel_cols_link4():
            n2 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n2))
                n2 += 1

        ## Formatting Sheet{Facility Level Summary}
        try:
            sheet.move_range("A1:I10000", rows=1)
        except:
            sheet.move_range("A1:I25000", rows=1)
            try:
                sheet.move_range("A1:I50000", rows=1)
            except:
                sheet.move_range("A1:I100000", rows=1)
            finally:
                sheet.move_range("A1:I1048576", rows=1)

        
        ## BOLD FONT IN FACILITY LEVEL SUMMARY 
        # for rows in sheet.iter_rows():
        #     for cell in rows:
        #         cell.font = Font(bold=True)

        ## ADDING HEADER IN Facility Level Summary
        sheet.oddHeader.center.text = "Facility Level Summary"
        sheet.oddHeader.center.size = 18
        sheet.oddHeader.center.font = "Tahoma,Bold"
        sheet.oddHeader.center.color = "CC3366"
        sheet.cell(row=1, column=3).value = 'Facility Level Summary'
        sheet.cell(row=2, column=1).value = 'Sr. No'
        sheet.cell(row=2, column=4).value = 'Facility Name'
        sheet['C1'].font = Font(size = 18, bold = True, color="003679")  

        ## EXPANDING ROWS OF SHEET (FACILITY LEVEL SUMMARY)
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 15
        sheet.column_dimensions['F'].width = 25
       
        cnt1, cnt2, cnt3, cnt4, cnt5 = 0, 0, 0, 0, 0
        for i in range(len(FList3)):
            if sheet.cell(row=i+3, column=8).value >= 50:
                cnt1 += 1
            elif sheet.cell(row=i+3, column=8).value < 50 and sheet.cell(row=i+3, column=8).value >= 25:
                cnt2 += 1
            elif sheet.cell(row=i+3, column=8).value < 25 and sheet.cell(row=i+3, column=8).value >= 10:
                cnt3 += 1
            elif sheet.cell(row=i+3, column=8).value < 10 and sheet.cell(row=i+3, column=8).value >= 5:
                cnt4 += 1
            elif sheet.cell(row=i+3, column=8).value < 5:
                cnt5 += 1

            for j in range(len(FList3[i])):
                '''
                # Create hyperlink to relevant cell
                '''
                link1 = "#'Checks giving Inconsistent'!B2"

                l1 = list(itertools.islice(excel_cols_link3(), len(FList3)+1))

                #update link
                link1 = link1.replace("B", l1[i+1])

                sheet.cell(row=i+3, column=5).hyperlink = link1
                sheet.cell(row=i+3, column=5).style = "Hyperlink"

                if sheet.cell(row=i+3, column=8).value >= 50:
                    sheet.cell(row=i+3, column=5).fill = PatternFill(fgColor='EF5350', fill_type = "solid") 
                elif sheet.cell(row=i+3, column=8).value < 50 and sheet.cell(row=i+3, column=8).value >= 25:
                    sheet.cell(row=i+3, column=5).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet.cell(row=i+3, column=8).value < 25 and sheet.cell(row=i+3, column=8).value >= 10:
                    sheet.cell(row=i+3, column=5).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet.cell(row=i+3, column=8).value < 10 and sheet.cell(row=i+3, column=8).value >= 5:
                    sheet.cell(row=i+3, column=5).fill = PatternFill(fgColor='00FF00', fill_type = "solid")
                elif sheet.cell(row=i+3, column=8).value < 5:
                    sheet.cell(row=i+3, column=5).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")

        cnt6, cnt7, cnt8, cnt9, cnt10 = 0, 0, 0, 0, 0
        for i in range(len(FList4)):
            if sheet.cell(row=i+3, column=9).value >= 50:
                cnt6 += 1
            elif sheet.cell(row=i+3, column=9).value < 50 and sheet.cell(row=i+3, column=9).value >= 25:
                cnt7 += 1
            elif sheet.cell(row=i+3, column=9).value < 25 and sheet.cell(row=i+3, column=9).value >= 10:
                cnt8 += 1
            elif sheet.cell(row=i+3, column=9).value < 10 and sheet.cell(row=i+3, column=9).value >= 5:
                cnt9 += 1
            elif sheet.cell(row=i+3, column=9).value < 5:
                cnt10 += 1

            for j in range(len(FList4[i])):

                '''
                # Create hyperlink to relevant cell
                '''
                link2 = "#'Checks giving PRE'!B2"

                l2 = list(itertools.islice(excel_cols_link4(), len(FList4)+1))
                
                #update link
                link2 = link2.replace("B", l2[i+1])

                sheet.cell(row=i+3, column=6).hyperlink = link2
                sheet.cell(row=i+3, column=6).style = "Hyperlink"

                if sheet.cell(row=i+3, column=9).value >= 50:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
                elif sheet.cell(row=i+3, column=9).value < 50 and sheet.cell(row=i+3, column=9).value >= 25:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet.cell(row=i+3, column=9).value < 25 and sheet.cell(row=i+3, column=9).value >= 10:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet.cell(row=i+3, column=9).value < 10 and sheet.cell(row=i+3, column=9).value >= 5:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='00FF00', fill_type = "solid")
                elif sheet.cell(row=i+3, column=9).value < 5:
                    sheet.cell(row=i+3, column=6).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")

        ## Bordering Facility Level summary Sheet
        for i in range(2, len(dataframeForSheet6)+2):
            sheet.cell(row=i, column=2).border = thick_border
            sheet.cell(row=i, column=3).border = thick_border
            sheet.cell(row=i, column=4).border = thick_border
            sheet.cell(row=i, column=5).border = thick_border
            sheet.cell(row=i, column=6).border = thick_border
            sheet.cell(row=i, column=7).border = thick_border


        # Coloring and palettes of Facility Guidance Sheet
        sheet['O4'] = 'Color Brackets'
        sheet['O5'].fill = PatternFill(fgColor="EF5350", fill_type = "solid")
        sheet['O6'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet['O7'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet['O8'].fill = PatternFill(fgColor="00FF00", fill_type = "solid")
        sheet['O9'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        sheet["P4"] = "Range"
        sheet["P5"] = ">= 50%"
        sheet["P6"] = "25 - 50%"
        sheet["P7"] = "10 - 25%"
        sheet["P8"] = "5 - 10%"
        sheet["P9"] = "< 5%"
        sheet["P10"] = "Total Facilities"

        sheet["Q4"] = "Inconsistent"
        sheet["Q5"] = cnt1
        sheet["Q6"] = cnt2
        sheet["Q7"] = cnt3
        sheet["Q8"] = cnt4
        sheet["Q9"] = cnt5
        sheet["Q10"] = cnt1 + cnt2 + cnt3 + cnt4 + cnt5

        sheet["R4"] = "Probable Reporting Error"
        sheet["R5"] = cnt6
        sheet["R6"] = cnt7
        sheet["R7"] = cnt8
        sheet["R8"] = cnt9
        sheet["R9"] = cnt10
        sheet["R10"] = cnt6 + cnt7 + cnt8 + cnt9 + cnt10


        ''' 
        GRAPH PLOTS
        '''
        
        Ranges = ['< 5%', '5 - 10%' , '10 - 25%', '25 - 50%', '>= 50%']
        Numbers_Inc = [cnt5, cnt4, cnt3, cnt2, cnt1]
        Numbers_PRE = [cnt10, cnt9, cnt8, cnt7, cnt6]
        
        # plotting a bar graph
        X_axis = np.arange(len(Ranges))

        plt.bar(X_axis - 0.2, Numbers_Inc, 0.4, label = 'Number of Inconsistents')
        plt.bar(X_axis + 0.2, Numbers_PRE, 0.4, label = 'Number of Probable Reporting Errors')
        # figure(figsize=(8, 8), dpi=50)
        plt.xticks(X_axis, Ranges)
        plt.xlabel("Condition")
        plt.ylabel("Number")
        plt.legend()
        plt.title('Error Summary Facility Wise')
        plt.savefig("myplot2.png", dpi = 80)

        img = openpyxl.drawing.image.Image('myplot2.png')
        img.anchor='M13'

        sheet.add_image(img)

        workbook.save(filename=filename)
        workbook.save(filename=filename)


        # Attention Required Sheet (Inconsistent)
        # =======================================

        workbook.active = sheet_2
        sheet_2.sheet_view.showGridLines = True

        sheet_2['A1'] = "Facility with Inconsistent"
        sheet_2.oddHeader.center.size = 18
        sheet_2.oddHeader.center.font = "Tahoma,Bold"
        sheet_2.oddHeader.center.color = "CC3366"
        sheet_2['A1'].font = Font(size = 18, bold = True, color="003679")

        ## Formatting Sheet_1{Validation Checkwise Summary}
        try:
            sheet_1.move_range("A1:G10000", rows=1)
        except:
            sheet_1.move_range("A1:G25000", rows=1)
            try:
                sheet_1.move_range("A1:G50000", rows=1)
            except:
                sheet_1.move_range("A1:G100000", rows=1)
            finally:
                sheet_1.move_range("A1:G1048576", rows=1)

        sheet_1.oddHeader.center.text = "Validation checkwise summary"
        sheet_1.cell(row=2, column=1).value = 'Sr. No'
        sheet_1.oddHeader.center.size = 14
        sheet_1.oddHeader.center.font = "Tahoma,Bold"
        sheet_1.oddHeader.center.color = "CC3366"
        sheet_1.cell(row=1, column=2).value = 'Validation checkwise summary'
        sheet_1['B1'].font = Font(size = 18, bold = True, color="003679")

        ## EXPANDING ROWS OF SHEET_1 (VALIDATION CHECKWISE SUMMARY)
        sheet_1.column_dimensions['B'].width = 20
        sheet_1.column_dimensions['C'].width = 20
        sheet_1.column_dimensions['D'].width = 15
        sheet_1.column_dimensions['E'].width = 25

        # HYPERLINKING FUNCTIONS
        def excel_cols_link1():
            n1 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n1))
                n1 += 1

        # HYPERLINKING FUNCTIONS
        def excel_cols_link2():
            n2 = 1
            while True:
                yield from (''.join(group) for group in itertools.product(string.ascii_uppercase, repeat=n2))
                n2 += 1
        
        cnt21, cnt22, cnt23, cnt24 = 0, 0, 0, 0
        for i in range(len(FList1)):
            if sheet_1.cell(row=i+3, column=6).value >= 25:
                cnt21 += 1
            elif sheet_1.cell(row=i+3, column=6).value < 25 and sheet_1.cell(row=i+3, column=6).value >= 10:
                cnt22 += 1
            elif sheet_1.cell(row=i+3, column=6).value < 10 and sheet_1.cell(row=i+3, column=6).value >= 5:
                cnt23 += 1
            elif sheet_1.cell(row=i+3, column=6).value < 5:
                cnt24 += 1

            for j in range(len(FList1[i])):
                sheet_2.cell(row=j+5,column=i+2).value = FList1[i][j]

                '''
                # Create hyperlink to relevant cell for sheet_1
                '''
                link1 = "#'Facility with Inconsistent'!B2"

                l1 = list(itertools.islice(excel_cols_link1(), len(FList1)+1))
                
                ## update link and then hyperlink it
                link1 = link1.replace("B", l1[i+1])
                sheet_1.cell(row=i+3, column=4).hyperlink = link1
                sheet_1.cell(row=i+3, column=4).style = "Hyperlink"

                if sheet_1.cell(row=i+3, column=6).value >= 25:
                    sheet_1.cell(row=i+3, column=4).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=6).value < 25 and sheet_1.cell(row=i+3, column=6).value >= 10:
                    sheet_1.cell(row=i+3, column=4).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=6).value < 10 and sheet_1.cell(row=i+3, column=6).value >= 5:
                    sheet_1.cell(row=i+3, column=4).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=6).value < 5:
                    sheet_1.cell(row=i+3, column=4).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")

            
            sheet_2['A2'].fill = PatternFill(fgColor="003679", fill_type = "solid")
            sheet_2['A3'].fill = PatternFill(fgColor="003679", fill_type = "solid")
            sheet_2['A4'].fill = PatternFill(fgColor="003679", fill_type = "solid")
            sheet_2['A5'].fill = PatternFill(fgColor="003679", fill_type = "solid")
            sheet_2['A2'].font = Font(color = "FFFFFF")
            sheet_2['A3'].font = Font(color = "FFFFFF")
            sheet_2['A4'].font = Font(color = "FFFFFF")
            sheet_2['A5'].font = Font(color = "FFFFFF")
            sheet_2.column_dimensions['A'].width = 35
            sheet_2.row_dimensions[2].height = 20
            sheet_2.row_dimensions[3].height = 20
            sheet_2.row_dimensions[4].height = 20
            sheet_2.row_dimensions[5].height = 20   
                
        
        workbook.save(filename=filename)


        # Attention Required Sheet (PRE)
        # =======================================
        workbook.active = sheet_3

        sheet_3.sheet_view.showGridLines = True
        sheet_3['A1'] = "Facility with PRE"
        sheet_3.oddHeader.center.size = 18
        sheet_3.oddHeader.center.font = "Tahoma,Bold"
        sheet_3.oddHeader.center.color = "CC3366"
        sheet_3['A1'].font = Font(size = 18, bold = True, color="003679")

        cnt25, cnt26, cnt27, cnt28 = 0, 0, 0, 0
        for i in range(len(FList2)):
            if sheet_1.cell(row=i+3, column=7).value >= 25:
                cnt25 += 1
            elif sheet_1.cell(row=i+3, column=7).value < 25 and sheet_1.cell(row=i+3, column=7).value >= 10:
                cnt26 += 1
            elif sheet_1.cell(row=i+3, column=7).value < 10 and sheet_1.cell(row=i+3, column=7).value >= 5:
                cnt27 += 1
            elif sheet_1.cell(row=i+3, column=7).value < 5:
                cnt28 += 1

            for j in range(len(FList2[i])):
                sheet_3.cell(row=j+5,column=i+2).value = FList2[i][j]
                sheet_3.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')

                '''
                # Create hyperlink to relevant cell for sheet_1
                '''
                link1 = "#'Facility with PRE'!B2"

                l1 = list(itertools.islice(excel_cols_link2(), len(FList2)+1))
                
                #update link
                link1 = link1.replace("B", l1[i+1])

                sheet_1.cell(row=i+3, column=5).hyperlink = link1
                sheet_1.cell(row=i+3, column=5).style = "Hyperlink"

                if sheet_1.cell(row=i+3, column=7).value >= 25:
                    sheet_1.cell(row=i+3, column=5).fill = PatternFill(fgColor='EF5350', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=7).value < 25 and sheet_1.cell(row=i+3, column=7).value >= 10:
                    sheet_1.cell(row=i+3, column=5).fill = PatternFill(fgColor='FFAF00', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=7).value < 10 and sheet_1.cell(row=i+3, column=7).value >= 5:
                    sheet_1.cell(row=i+3, column=5).fill = PatternFill(fgColor='C0C000', fill_type = "solid")
                elif sheet_1.cell(row=i+3, column=7).value < 5:
                    sheet_1.cell(row=i+3, column=5).fill = PatternFill(fgColor='00AF5F', fill_type = "solid")

                
            # Colors
            for k in range(1, len(FList2[i])+100):
                sheet_3.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                # sheet_3.cell(row=2, column=i+2).fill = PatternFill(fgColor="fff5be", fill_type = "solid")

        sheet_3['A2'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_3['A3'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_3['A4'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_3['A5'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_3['A2'].font = Font(color = "FFFFFF")
        sheet_3['A3'].font = Font(color = "FFFFFF")
        sheet_3['A4'].font = Font(color = "FFFFFF")
        sheet_3['A5'].font = Font(color = "FFFFFF")
        sheet_3.column_dimensions['A'].width = 35
        sheet_3.row_dimensions[2].height = 20
        sheet_3.row_dimensions[3].height = 20
        sheet_3.row_dimensions[4].height = 20
        sheet_3.row_dimensions[5].height = 20   

        # Coloring of Validation Summary Sheet
        sheet_1['J10'] = "Color Brackets"
        sheet_1['J11'].fill = PatternFill(fgColor="EF5350", fill_type = "solid")
        sheet_1['J12'].fill = PatternFill(fgColor="FFAF00", fill_type = "solid")
        sheet_1['J13'].fill = PatternFill(fgColor="C0C000", fill_type = "solid")
        sheet_1['J14'].fill = PatternFill(fgColor="00AF5F", fill_type = "solid")

        sheet_1['K10'] = "Range"
        sheet_1['K11'] = ">= 25%"
        sheet_1['K12'] = "10 - 25%"
        sheet_1['K13'] = "5 - 10%"
        sheet_1['K14'] = "< 5%"
        sheet_1['K15'] = "Total Indicators"

        sheet_1['L10'] = "Inconsistent"
        sheet_1['L11'] = cnt21
        sheet_1['L12'] = cnt22
        sheet_1['L13'] = cnt23
        sheet_1['L14'] = cnt24
        sheet_1['L15'] = cnt21 + cnt22 + cnt23 + cnt24

        sheet_1['M10'] = "Probable Reporting Error"
        sheet_1['M11'] = cnt25
        sheet_1['M12'] = cnt26
        sheet_1['M13'] = cnt27
        sheet_1['M14'] = cnt28
        sheet_1['M15'] = cnt25 + cnt26 + cnt27 + cnt28

        ## Bordering Validation checkwise summary Sheet
        for i in range(2, len(dataframeForSheet4)+2):
            sheet_1.cell(row=i, column=2).border = thick_border
            sheet_1.cell(row=i, column=3).border = thick_border
            sheet_1.cell(row=i, column=4).border = thick_border
            sheet_1.cell(row=i, column=5).border = thick_border
        
        workbook.save(filename=filename)


        # Checks Sheet (Inconsistent)
        # =======================================
        sheet_4['A1'] = "Checks giving Inconsistent"
        sheet_4.oddHeader.center.size = 18
        sheet_4.oddHeader.center.font = "Tahoma,Bold"
        sheet_4.oddHeader.center.color = "CC3366"
        sheet_4['A1'].font = Font(size = 18, bold = True, color="003679")

        sheet_4.sheet_view.showGridLines = True
        for i in range(len(FList3)):

            for j in range(len(FList3[i])):
                sheet_4.cell(row=j+4,column=i+2).value = FList3[i][j]

            # Colors
            for k in range(1, len(FList3[i])+5):
                sheet_4.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')

        sheet_4['A2'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_4['A3'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_4['A4'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_4['A2'].font = Font(color = "FFFFFF")
        sheet_4['A3'].font = Font(color = "FFFFFF")
        sheet_4['A4'].font = Font(color = "FFFFFF")
        # set the width of the row
        sheet_4.column_dimensions['A'].width = 36
        sheet_4.row_dimensions[2].height = 20
        sheet_4.row_dimensions[3].height = 20
        sheet_4.row_dimensions[4].height = 20

        # ## Bordering Checks Giving Inconsistent Sheet
        # for i in range(1, len(dataframeForSheet6)+50):
        #     sheet_4.cell(row=1, column=i).border = thick_border
        #     sheet_4.cell(row=2, column=i).border = thick_border
        #     sheet_4.cell(row=1, column=i).border = thick_border

        workbook.save(filename=filename)


        # Checks Sheet (PRE)
        # =======================================
        workbook.active = sheet_5

        sheet_5['A1'] = "Checks giving PRE"
        sheet_5.oddHeader.center.size = 18
        sheet_5.oddHeader.center.font = "Tahoma,Bold"
        sheet_5.oddHeader.center.color = "CC3366"
        sheet_5['A1'].font = Font(size = 18, bold = True, color="003679")

        sheet_5.sheet_view.showGridLines = True
        for i in range(len(FList4)):
            for j in range(len(FList4[i])):
                sheet_5.cell(row=j+4,column=i+2).value = FList4[i][j]

            # Colors
            for k in range(1, len(FList4[i])+5):
                sheet_5.cell(row=4, column=i+1).alignment = Alignment(horizontal='center')
                # sheet_5.cell(row=2, column=i+2).fill = PatternFill(fgColor="fff5be", fill_type = "solid")
                # Top 10 Facility Checks not working well
                # sheet_5.cell(row=k+2, column=i+2).fill = PatternFill(fgColor="EF9A9A", fill_type = "solid")

        sheet_5['A2'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_5['A3'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_5['A4'].fill = PatternFill(fgColor="003679", fill_type = "solid")
        sheet_5['A2'].font = Font(color = "FFFFFF")
        sheet_5['A3'].font = Font(color = "FFFFFF")
        sheet_5['A4'].font = Font(color = "FFFFFF")
        sheet_5.column_dimensions['A'].width = 25
        sheet_5.row_dimensions[2].height = 20
        sheet_5.row_dimensions[3].height = 20
        sheet_5.row_dimensions[4].height = 20 

        workbook.save(filename=filename)

        # Coloring validated data sheet
        workbook.active = sheet_6

        for i in range(len(df.axes[0])):
            for j in range(len(df.axes[1])):
                pattern_ = re.compile("^I/i")
                if pattern_.match(str(sheet_6.cell(row=i+2, column=j+2).value)):
                    sheet_6.cell(row=i+2, column=j+2).fill = PatternFill(fgColor="F44336", fill_type = "solid")

        
        workbook.save(filename = filename)


        workbook.active = sheet_0

        ## Bordering Description Sheet
        for i in range(4, 13):
            sheet_0.cell(row=i+1, column=2).border = thick_border
            sheet_0.cell(row=i+1, column=3).border = thick_border

        for i in range(16, 20):
            sheet_0.cell(row=i+1, column=2).border = thick_border
            sheet_0.cell(row=i+1, column=3).border = thick_border


        ## Bordering Validation checkwise summary Sheet
        for i in range(2, len(dataframeForSheet4)+2):
            sheet_1.cell(row=i, column=2).border = thick_border
            sheet_1.cell(row=i, column=3).border = thick_border
            sheet_1.cell(row=i, column=4).border = thick_border
            sheet_1.cell(row=i, column=5).border = thick_border


        sheet_0.sheet_view.showGridLines = False
        sheet_0.merge_cells('A1:B2')
        cell_A1 = sheet_0.cell(row= 1, column= 1)
        cell_A1.value = 'Data Validation Check Tool:'  
        cell_A1.alignment = Alignment(horizontal='center', vertical='center', indent=0) 
        cell_A1.fill = PatternFill("solid", fgColor="00003366")
        cell_A1.font = Font(color="00FFFFFF", size = 14, bold = True)

        sheet_0.column_dimensions['A'].width = 15
        sheet_0.column_dimensions['B'].width = 30
        sheet_0.column_dimensions['C'].width = 80
        sheet_0.column_dimensions['D'].width = 80

        sheet_0.cell(row=3, column=3).alignment = Alignment(horizontal='right')
        sheet_0['B4'] = "Summary Sheet Guide:"
        sheet_0['B5'] = "Facility Type: " + self.lineEdit_2.text()
        sheet_0['B6'] = "Description"
        sheet_0['B7'] = "Facility Level Summary"
        sheet_0['B8'] = "Validation checkwise summary"
        sheet_0['B9'] = "Facility with Inconsistent"
        sheet_0['B10'] = "Facility with PRE"
        sheet_0['B11'] = "Checks giving Inconsistent"
        sheet_0['B12'] = "Checks giving PRE"
        sheet_0['B13'] = "Validated Data"
        sheet_0['B16'] = "Validation Check Outcome Definition : "
        sheet_0['B17'] = "Consistent"
        sheet_0['B18'] = "Inconsistent"
        sheet_0['B19'] = "Probable reporting error"
        sheet_0['B20'] = "Blank"

        sheet_0['C4'] = ""
        sheet_0['C5'] = "Duration: " + self.lineEdit_3.text()
        sheet_0['C6'] = "Description of sheets, important terminologies and other explanations"
        sheet_0['C7'] = "This sheet gives the counts of errors corresponding to each facility name. The colour coding is \n done as per the buckets, considering the percentage of the number of inconsistent/PRE out of\n the total validation checks in that facility type for each facility name. This is also shown \ngraphically below the bucketing table. Clicking on the figures of the original table, the user will be directed to the ???Checks giving inconsistent???\n or ???Checks giving PRE??? tabs, showing the validation checks for which, the errors creeped in."
        sheet_0['C8'] = "This sheet gives us the count of the facilities giving inconsistent and probable reporting error\n within the datasets. The colour coding is done as per the buckets, which were created considering the percentage of the\n number of facility names out of all the facilities who have reported inconsistent/PRE. This is also shown graphically below the\n bucketing table. Clicking on the figures of the original table, the user will be directed to the ???Facility with inconsistent??? \nor ???Facility with PRE??? tabs, showing the facilities for which, the errors creeped in."
        sheet_0['C9'] = "This sheet provides the list of the facilities giving Inconsistencies within the dataset\n with the count and the condition and the description of the condition."
        sheet_0['C10'] = "This sheet provides the list of the facilities giving probable reporting error within the dataset with the\n count and the condition and the description of the condition."
        sheet_0['C11'] = "This sheet provides the list of the checks giving Inconsistencies within the dataset with the count and the\n description."
        sheet_0['C12'] = "This sheet provides the list of the checks giving Probable Reporting Error within the dataset with the count\n and the description"
        sheet_0['C13'] = "This sheet is the complete raw data with the checks embedded with the sheet which gives a detailed\n information, row wise identification of the inconsistent and PRE with \nthe problematic indicators because of which the user is getting errors in its dataset"
        sheet_0['C16'] = ""
        sheet_0['C17'] = "The validation check holds true and needs no scrutiny."
        sheet_0['C18'] = "The validation check fails and the inconsistent data item is flagged. "
        sheet_0['C19'] = "The validation check may fail and it is subject to confirmation with the concerned authority. Check and verify."
        sheet_0['C20'] = "The validation where all the items are blank."

        sheet_0['B4'].font = Font(size = 14, bold = True)
        sheet_0['B16'].font = Font(size = 14, bold = True)
        sheet_0['B4'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['C4'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['B16'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['C16'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        # sheet_0['C4'].fill = PatternFill(fgColor="FFFFCC", fill_type = "solid")
        sheet_0['B5'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['B5'].font = Font(bold = True)
        sheet_0['C5'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['C5'].font = Font(bold = True)

        for i in range(16, 20):
            sheet_0.cell(row=i+1, column=2).border = thick_border
            sheet_0.cell(row=i+1, column=3).border = thick_border

        sheet_0['B23'] = "Examples"
        sheet_0['B23'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")
        sheet_0['B23'].font = Font(size = 14, bold = True)
        sheet_0['C23'].fill = PatternFill(fgColor="00C0C0C0", fill_type = "solid")

        #First 
        sheet_0['B24'] = "Related data items (???parent-child relation???):"
        sheet_0['B24'].font = Font(size = 12, bold = True)

        sheet_0.merge_cells('B25:B29')  
  
        cell_B25 = sheet_0.cell(row= 25, column= 2)  
        cell_B25.value = '1.1.1 Out of the total ANC registered, number registered within 1st trimester (within 12 weeks)'  
        cell_B25.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_B25.fill = PatternFill("solid", fgColor="00003366")
        cell_B25.font = Font(color="00FFFFFF")
        

        sheet_0.merge_cells('C25:C29')  
  
        cell_C25 = sheet_0.cell(row= 25, column= 3)  
        cell_C25.value = '1.1 Total number of pregnant women registered for ANC'  
        cell_C25.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_C25.fill = PatternFill("solid", fgColor="00003366")
        cell_C25.font = Font(color="00FFFFFF")
        

        sheet_0.merge_cells('D25:D29')  
  
        cell_D25 = sheet_0.cell(row= 25, column= 4)  
        cell_D25.value = 'Outcome (for 1.1.1 <= 1.1)'  
        cell_D25.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_D25.fill = PatternFill("solid", fgColor="00003366")
        cell_D25.font = Font(color="00FFFFFF")
        
        
        sheet_0['B30'] = 'Value'
        sheet_0['B30'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B30'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B30'].font = Font(color="00FFFFFF")
        sheet_0['B30'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))
        

        sheet_0['B31'] = 'Value'
        sheet_0['B31'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B31'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B31'].font = Font(color="00FFFFFF")
        sheet_0['B31'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B32'] = 'Null'
        sheet_0['B32'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B32'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B32'].font = Font(color="00FFFFFF")
        sheet_0['B32'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B33'] = 'Value'
        sheet_0['B33'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B33'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B33'].font = Font(color="00FFFFFF")
        sheet_0['B33'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B34'] = 'Blank'
        sheet_0['B34'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B34'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B34'].font = Font(color="00FFFFFF")
        sheet_0['B34'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C30'] = 'Value'
        sheet_0['C30'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C30'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C30'].font = Font(color="00000000", bold = False)
        sheet_0['C30'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C31'] = 'Value'
        sheet_0['C31'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C31'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C31'].font = Font(color="00000000", bold = False)
        sheet_0['C31'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C32'] = 'Value'
        sheet_0['C32'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C32'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C32'].font = Font(color="00000000", bold = False)
        sheet_0['C32'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C33'] = 'Null'
        sheet_0['C33'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C33'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C33'].font = Font(color="00000000", bold = False)
        sheet_0['C33'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C34'] = 'Blank'
        sheet_0['C34'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C34'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C34'].font = Font(color="00000000", bold = False)
        sheet_0['C34'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D30'] = 'Consistent'
        sheet_0['D30'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D30'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D30'].font = Font(color="00000000", bold = False)
        sheet_0['D30'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D31'] = 'Inconsistent (when condition fails)'
        sheet_0['D31'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D31'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D31'].font = Font(color="00000000", bold = False)
        sheet_0['D31'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D32'] = 'Probable Reporting Error'
        sheet_0['D32'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D32'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D32'].font = Font(color="00000000", bold = False)
        sheet_0['D32'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D33'] = 'Inconsistent'
        sheet_0['D33'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D33'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D33'].font = Font(color="00000000", bold = False)
        sheet_0['D33'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D34'] = 'Blank'
        sheet_0['D34'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D34'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D34'].font = Font(color="00000000", bold = False)
        sheet_0['D34'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))


        sheet_0.row_dimensions[30].height = 15
        sheet_0.row_dimensions[31].height = 15
        sheet_0.row_dimensions[32].height = 15
        sheet_0.row_dimensions[33].height = 15
        sheet_0.row_dimensions[34].height = 15


        #Second
        sheet_0['B37'] = "Complex check: involving multiple data items in either side of the check."
        sheet_0['B37'].font = Font(size = 12, bold = True)

        sheet_0.merge_cells('B38:B42')  
  
        cell_B38 = sheet_0.cell(row= 38, column= 2)  
        cell_B38.value = '14.3.3 Number of Left Against Medical Advice (LAMA) cases'  
        cell_B38.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_B38.fill = PatternFill("solid", fgColor="00003366")
        cell_B38.font = Font(color="00FFFFFF")

        sheet_0.merge_cells('C38:C42')  
  
        cell_C38 = sheet_0.cell(row= 38, column= 3)  
        cell_C38.value = '14.3.1.a + 14.3.1.b + 14.3.2.a + 14.3.2.b  Inpatient(Male)-Children<18yrs + Inpatient(Male)-Adults + Inpatient(Female)-Children<18yrs + Inpatient(Female)-Adults'  
        cell_C38.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_C38.fill = PatternFill("solid", fgColor="00003366")
        cell_C38.font = Font(color="00FFFFFF")

        sheet_0.merge_cells('D38:D42')  
  
        cell_D38 = sheet_0.cell(row= 38, column= 4)  
        cell_D38.value = 'Outcome (for 14.3.3 <= 14.3.1.a + 14.3.1.b + 14.3.2.a + 14.3.2.b)'  
        cell_D38.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_D38.fill = PatternFill("solid", fgColor="00003366")
        cell_D38.font = Font(color="00FFFFFF")

        sheet_0['B43'] = 'Value'
        sheet_0['B43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B43'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B43'].font = Font(color="00FFFFFF")
        sheet_0['B43'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B44'] = 'Value'
        sheet_0['B44'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B44'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B44'].font = Font(color="00FFFFFF")
        sheet_0['B44'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B45'] = 'Null'
        sheet_0['B45'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B45'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B45'].font = Font(color="00FFFFFF")
        sheet_0['B45'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B46'] = 'Value'
        sheet_0['B46'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B46'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B46'].font = Font(color="00FFFFFF")
        sheet_0['B46'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B47'] = 'Blank'
        sheet_0['B47'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B47'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B47'].font = Font(color="00FFFFFF")
        sheet_0['B47'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C43'] = 'Value'
        sheet_0['C43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C43'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C43'].font = Font(color="00000000", bold = False)
        sheet_0['C43'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C44'] = 'Value'
        sheet_0['C44'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C44'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C44'].font = Font(color="00000000", bold = False)
        sheet_0['C44'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C45'] = 'Value'
        sheet_0['C45'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C45'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C45'].font = Font(color="00000000", bold = False)
        sheet_0['C45'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C46'] = 'Null'
        sheet_0['C46'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C46'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C46'].font = Font(color="00000000", bold = False)
        sheet_0['C46'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C47'] = 'Blank'
        sheet_0['C47'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C47'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C47'].font = Font(color="00000000", bold = False)
        sheet_0['C47'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D43'] = 'Consistent'
        sheet_0['D43'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D43'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D43'].font = Font(color="00000000", bold = False)
        sheet_0['D43'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D44'] = 'Inconsistent (when condition fails)'
        sheet_0['D44'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D44'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D44'].font = Font(color="00000000", bold = False)
        sheet_0['D44'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D45'] = 'Probable Reporting Error'
        sheet_0['D45'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D45'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D45'].font = Font(color="00000000", bold = False)
        sheet_0['D45'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D46'] = 'Inconsistent'
        sheet_0['D46'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D46'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D46'].font = Font(color="00000000", bold = False)
        sheet_0['D46'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D47'] = 'Blank'
        sheet_0['D47'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D47'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D47'].font = Font(color="00000000", bold = False)
        sheet_0['D47'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))


        sheet_0.row_dimensions[43].height = 15
        sheet_0.row_dimensions[44].height = 15
        sheet_0.row_dimensions[45].height = 15
        sheet_0.row_dimensions[46].height = 15
        sheet_0.row_dimensions[47].height = 15

        #Third
        sheet_0['B50'] = "Recurring data items [service for one data items may be provided over months]:  Considering permissible limit of +-50% i.e., if the disparity in the \ntwo data items is more than the limit then it is a probable reporting error."
        sheet_0['B50'].font = Font(size = 12, bold = True)
        
        sheet_0.merge_cells('B51:B55')  
  
        cell_B50 = sheet_0.cell(row= 51, column= 2)  
        cell_B50.value = '1.2.4 Number of PW given 180 Iron Folic Acid (IFA) tablets'  
        cell_B50.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_B50.fill = PatternFill("solid", fgColor="00003366")
        cell_B50.font = Font(color="00FFFFFF")

        sheet_0.merge_cells('C51:C55')  
  
        cell_C50 = sheet_0.cell(row= 51, column= 3)  
        cell_C50.value = '1.1 Total number of pregnant women registered for ANC'  
        cell_C50.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_C50.fill = PatternFill("solid", fgColor="00003366")
        cell_C50.font = Font(color="00FFFFFF")

        sheet_0.merge_cells('D51:D55')  
  
        cell_D50 = sheet_0.cell(row= 51, column= 4)  
        cell_D50.value = 'Outcome (for 1.2.4 <= 1.1)'  
        cell_D50.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0) 
        cell_D50.fill = PatternFill("solid", fgColor="00003366")
        cell_D50.font = Font(color="00FFFFFF")

        sheet_0['B56'] = 'Value'
        sheet_0['B56'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B56'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B56'].font = Font(color="00FFFFFF")
        sheet_0['B56'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B57'] = 'Value'
        sheet_0['B57'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B57'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B57'].font = Font(color="00FFFFFF")
        sheet_0['B57'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B58'] = 'Null'
        sheet_0['B58'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B58'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B58'].font = Font(color="00FFFFFF")
        sheet_0['B58'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B59'] = 'Value'
        sheet_0['B59'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B59'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B59'].font = Font(color="00FFFFFF")
        sheet_0['B59'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['B60'] = 'Blank'
        sheet_0['B60'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['B60'].fill = PatternFill("solid", fgColor="00003366")
        sheet_0['B60'].font = Font(color="00FFFFFF")
        sheet_0['B60'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C56'] = 'Value'
        sheet_0['C56'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C56'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C56'].font = Font(color="00000000", bold = False)
        sheet_0['C56'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C57'] = 'Value'
        sheet_0['C57'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C57'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C57'].font = Font(color="00000000", bold = False)
        sheet_0['C57'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C58'] = 'Value'
        sheet_0['C58'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C58'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C58'].font = Font(color="00000000", bold = False)
        sheet_0['C58'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C59'] = 'Null'
        sheet_0['C59'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C59'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['C59'].font = Font(color="00000000", bold = False)
        sheet_0['C59'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['C60'] = 'Blank'
        sheet_0['C60'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['C60'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['C60'].font = Font(color="00000000", bold = False)
        sheet_0['C60'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D56'] = 'Consistent [Probable Reporting Error (if LHS is less than 50% of RHS)]'
        sheet_0['D56'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D56'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D56'].font = Font(color="00000000", bold = False)
        sheet_0['D56'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D57'] = 'Inconsistent (when condition fails) [Probable Reporting Error (is LHS is more than 50% of RHS)]'
        sheet_0['D57'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D57'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D57'].font = Font(color="00000000", bold = False)
        sheet_0['D57'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D58'] = 'Probable Reporting Error'
        sheet_0['D58'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D58'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D58'].font = Font(color="00000000", bold = False)
        sheet_0['D58'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D59'] = 'Probable Reporting Error'
        sheet_0['D59'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D59'].fill = PatternFill("solid", fgColor="00FFFFFF")
        sheet_0['D59'].font = Font(color="00000000", bold = False)
        sheet_0['D59'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0['D60'] = 'Blank'
        sheet_0['D60'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True, indent=0)
        sheet_0['D60'].fill = PatternFill("solid", fgColor="00C0C0C0")
        sheet_0['D60'].font = Font(color="00000000", bold = False)
        sheet_0['D60'].border = Border(top = Side(style='thin'), left= Side(style='thin'), right= Side(style='thin'), bottom= Side(style='thin'))

        sheet_0.row_dimensions[56].height = 15
        sheet_0.row_dimensions[57].height = 15
        sheet_0.row_dimensions[58].height = 15
        sheet_0.row_dimensions[59].height = 15
        sheet_0.row_dimensions[60].height = 15

        ## WRAPPING TEXT IN DESCRIPTION SHEET
        for rows in sheet_0.iter_rows():
            for cell in rows:
                cell.alignment = Alignment(wrapText=True)
        

        # delete useless columns of Facility Level Summary
        sheet.delete_cols(8, 2)

        # delete useless columns of sheet_1
        sheet_1.delete_cols(6, 2)
        workbook.save(filename=filename)



        self.popup.close()

        # Create the messagebox object
        self.msg = QMessageBox()
        # Set the information icon
        self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
        self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
        # Set the main message
        self.msg.setText("Excel file downloaded in the selected location \n\n ?????????????????? ??????????????? ??????????????? ??????????????? ?????? ????????????????????? ?????? ?????????")
        # Set the title of the window
        self.msg.setWindowTitle(" ")
        # Display the message box
        self.msg.show()

        os.remove("FileName.csv")
        os.remove("myplot2.png")


    # Reset
    def reset(self):
        QtCore.QCoreApplication.quit()
        status = QtCore.QProcess.startDetached(sys.executable, sys.argv)
      
        print('Please WAIT the data and settings are resetting ..... ')
        print('After resetting TOOL will open again .....')


    # # RESET FUNCTION
    # def reset(self):
    #     self.lineEdit.clear()
    #     self.lineEdit_2.clear()
    #     self.lineEdit_3.clear()
    #     self.pushButton.setEnabled(True)
    #     self.pushButton_2.setEnabled(True)
    #     self.pushButton_3.setEnabled(True)
    #     self.pushButton_4.setEnabled(True)
    #     self.pushButton_5.setEnabled(True)
    #     self.pushButton_6.setEnabled(True)
    #     self.pushButton_7.setEnabled(True)
    #     self.pushButton_3.setText('-- All Selected --')
    #     self.pushButton_4.setText('-- All Selected --')
    #     self.pushButton_5.setText('-- All Selected --')
    #     self.pushButton_6.setText('-- All Selected --')


    #     # Create the messagebox object
    #     self.msg = QMessageBox()
    #     # Set the information icon
    #     self.msg.setWindowIcon(QtGui.QIcon('checked.png'))
    #     self.msg.setStyleSheet("QLabel { margin-right: 15px ; font-size: 18px; font-family: Arial;} QPushButton {background-color:lightgreen; font-family: Arial; font-size:20px;} ")
    #     # Set the main message
    #     self.msg.setText("Reset Successful , Now you can upload data again.")
    #     # Set the title of the window
    #     self.msg.setWindowTitle("Reset Successful Message")
    #     # Display the message box
    #     self.msg.show()

    # Display methodology pdf in browser
    def guidelines(self):
        os.system('start guidelines.pdf')

    # Display methodology pdf in browser
    def UserManual(self):
        os.system('start UserManual.pdf')


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    TabWidget = QtWidgets.QTabWidget()
    ui = Ui_TabWidget()
    ui.setupUi(TabWidget)
    TabWidget.show()
    sys.exit(app.exec_())
